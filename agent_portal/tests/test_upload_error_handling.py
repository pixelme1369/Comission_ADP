"""Covers two related fixes for the reported "Internal Server Error" on
upload:

1. None of the three upload routes (CRM, Cordoba payout, commission
   history) had any exception handling — an unexpected failure (a database
   constraint violation, a malformed row the parser didn't anticipate,
   anything) bubbled all the way up as a raw Flask 500 page with no
   indication of what broke or that nothing was saved. Each now catches,
   rolls back, and flashes a clear error instead.

2. The actual root cause reproduced: several client_record VARCHAR columns
   (crm_id, status, stage, the date columns, pay_freq) were sized off a
   single sample CRM row rather than any real limit. A real CRM export with
   a longer value in one of them raised
   psycopg2.errors.StringDataRightTruncation ("value too long for type
   character varying(N)") — verified directly against a real local
   Postgres database via the actual /admin/upload-csv route (SQLite,
   used by these tests, does not enforce VARCHAR length at all, so that
   specific failure can't be reproduced here — see the PR/commit
   description for that transcript). The columns are now widened, with a
   self-serve "Fix Now" migration (mirroring
   run_nullable_password_migration) for databases created before this
   change, tested the same way as test_admin_migration.py."""

import io

from agent_portal.models import Agent


def _make_admin(db, email="admin@company.com"):
    agent = Agent(email=email, display_name="Admin User", is_admin=True)
    agent.set_password("pw12345")
    db.session.add(agent)
    db.session.commit()
    return agent


def _login_as(client, agent):
    with client.session_transaction() as sess:
        sess["_user_id"] = str(agent.id)
        sess["_fresh"] = True


CSV_HEADERS = [
    "ID", "Sales Rep", "Full Name", "1st Payment Cleared Date", "Dropped Date",
    "Status", "Enrolled Debt", "# NSF", "Payments Made", "Pay Freq.",
]


def _csv_bytes(rows):
    import csv as csv_module
    out = io.StringIO()
    writer = csv_module.DictWriter(out, fieldnames=CSV_HEADERS)
    writer.writeheader()
    for r in rows:
        writer.writerow({h: r.get(h, "") for h in CSV_HEADERS})
    return out.getvalue().encode("utf-8")


class TestUploadRoutesDoNotCrashOnUnexpectedErrors:
    def test_csv_upload_flashes_error_instead_of_500(self, app, db, client, monkeypatch):
        admin = _make_admin(db)
        _login_as(client, admin)

        def boom(*a, **kw):
            raise Exception('value too long for type character varying(50)')
        monkeypatch.setattr("agent_portal.routes_admin.db.session.commit", boom)

        csv_bytes = _csv_bytes([{
            "ID": "1", "Sales Rep": "Agent A", "Full Name": "Client A",
            "1st Payment Cleared Date": "04/10/2026", "Status": "Active",
            "Enrolled Debt": "10000", "# NSF": "0", "Payments Made": "1", "Pay Freq.": "Monthly",
        }])
        resp = client.post(
            "/admin/upload-csv",
            data={"csv_file": (io.BytesIO(csv_bytes), "crm.csv")},
            content_type="multipart/form-data", follow_redirects=True,
        )
        assert resp.status_code == 200
        assert b"Import failed" in resp.data
        assert b"nothing was saved" in resp.data
        assert b"Internal Server Error" not in resp.data

    def test_commission_history_upload_flashes_error_instead_of_500(self, app, db, client, monkeypatch):
        admin = _make_admin(db)
        _login_as(client, admin)

        def boom(*a, **kw):
            raise Exception("simulated failure")
        monkeypatch.setattr("agent_portal.routes_admin.import_commission_history_files", boom)

        history_bytes = "Month,ID,Sales Rep,Enrolled Debt,To subtract\nApril,1,Agent A,10000,\n".encode("utf-8")
        resp = client.post(
            "/admin/upload-commission-history",
            data={"history_year": "2026", "history_file": (io.BytesIO(history_bytes), "history.csv")},
            content_type="multipart/form-data", follow_redirects=True,
        )
        assert resp.status_code == 200
        assert b"Commission history import failed" in resp.data
        assert b"Internal Server Error" not in resp.data

    def test_cordoba_upload_flashes_error_instead_of_500(self, app, db, client, monkeypatch):
        admin = _make_admin(db)
        _login_as(client, admin)

        def boom(*a, **kw):
            raise Exception("simulated failure")
        monkeypatch.setattr("agent_portal.routes_admin.process_cordoba_file", boom)

        resp = client.post(
            "/admin/upload-cordoba-payout",
            data={"cordoba_file": (io.BytesIO(b"not a real xlsx"), "cordoba.xlsx")},
            content_type="multipart/form-data", follow_redirects=True,
        )
        assert resp.status_code == 200
        assert b"Cordoba payout processing failed" in resp.data
        assert b"Internal Server Error" not in resp.data

    def test_cordoba_upload_returns_a_response_on_success(self, app, db, client):
        """Regression test: upload_cordoba_payout's success path used to fall
        off the end of the function with no return statement -- Flask raised
        TypeError('the view function did not return a valid response') on
        EVERY successful Cordoba payout upload (the data was still committed
        inside process_cordoba_file, only the response was missing), so a
        real admin saw a raw 500 instead of the success flash messages. A
        minimal, valid, empty-of-matches workbook exercises the success path
        without touching process_cordoba_file at all."""
        admin = _make_admin(db)
        _login_as(client, admin)

        import openpyxl
        wb = openpyxl.Workbook()
        first_pays = wb.active
        first_pays.title = "First Pays"
        first_pays.append([
            "Assigned Marketing", "Enrolled Date", "ID", "Full Name", "Status",
            "Enrolled Debt", "1st Payment Cleared Date", "Payments Made",
            "Marketing Payment Cleared", "Home Phone", "Credit Score", "Source File",
        ])
        epf = wb.create_sheet("EPF")
        epf.append([
            "Contact ID", "Enrolled Date", "Enrolled Debt", "Marketing Company",
            "Full Name", "Amount", "Cleared Date", "Amount Owed", "Source File",
        ])
        chargebacks = wb.create_sheet("Chargebacks")
        chargebacks.append([
            "Assigned Company", "Enrolled Date", "ID", "Full Name", "Status",
            "Marketing Payout Debt", "1st Payment Cleared Date", "Pay Freq.",
            "Payments Made", "Marketing Payment Cleared", "Marketing Payment Chargeback",
            "Dropped Date", "Source File",
        ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = client.post(
            "/admin/upload-cordoba-payout",
            data={"cordoba_file": (buf, "cordoba.xlsx")},
            content_type="multipart/form-data", follow_redirects=True,
        )
        assert resp.status_code == 200
        assert b"Cordoba payout processed" in resp.data
        assert b"Internal Server Error" not in resp.data


class TestColumnWidenFixNowCardVisibility:
    def test_card_hidden_on_a_fresh_database(self, app, db, client):
        # conftest's fresh SQLite db is built from the current (widened) model.
        admin = _make_admin(db)
        _login_as(client, admin)
        resp = client.get("/admin/")
        assert b"CRM Import Fix Needed" not in resp.data

    def test_card_shown_when_introspection_reports_narrow_columns(self, app, db, client, monkeypatch):
        admin = _make_admin(db)
        _login_as(client, admin)
        monkeypatch.setattr(
            "agent_portal.routes_admin._client_record_columns_are_wide_enough", lambda: False,
        )
        resp = client.get("/admin/")
        assert b"CRM Import Fix Needed" in resp.data
        assert b"Fix Now" in resp.data

    def test_detection_fails_open_on_introspection_error(self, app, db, monkeypatch):
        from agent_portal import routes_admin

        def boom(*a, **kw):
            raise RuntimeError("no introspection access")
        monkeypatch.setattr(routes_admin, "sa_inspect", boom)
        assert routes_admin._client_record_columns_are_wide_enough() is True


class TestColumnWidenMigrationRoute:
    def test_successful_migration_flashes_success(self, app, db, client, monkeypatch):
        admin = _make_admin(db)
        _login_as(client, admin)
        monkeypatch.setattr("agent_portal.db.session.execute", lambda *a, **kw: None)
        resp = client.post("/admin/migrate/widen-client-record-columns", follow_redirects=True)
        assert resp.status_code == 200
        assert b"Migration applied" in resp.data

    def test_failed_migration_flashes_error_without_crashing(self, app, db, client, monkeypatch):
        admin = _make_admin(db)
        _login_as(client, admin)

        def boom(*a, **kw):
            raise Exception("ALTER COLUMN is not supported on this dialect")
        monkeypatch.setattr("agent_portal.db.session.execute", boom)
        resp = client.post("/admin/migrate/widen-client-record-columns", follow_redirects=True)
        assert resp.status_code == 200
        assert b"Migration failed" in resp.data

    def test_non_admin_cannot_run_migration(self, app, db, client):
        agent = Agent(email="agent@company.com", display_name="Regular Agent", is_admin=False)
        agent.set_password("pw12345")
        db.session.add(agent)
        db.session.commit()
        _login_as(client, agent)
        resp = client.post("/admin/migrate/widen-client-record-columns", follow_redirects=True)
        assert resp.status_code == 200
        assert b"Admin access required" in resp.data
