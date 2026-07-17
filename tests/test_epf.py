"""EPF tab behavior (owner decision, July 2026): each row credits ONE UNIT toward
the matched agent's tier for the month, but never adds its own debt.

EPF rows are matched by Contact ID to our ClientRecord history to find the sales
rep, placed in the month of the EPF tab's Cleared Date, and shown in an "EPF"
section on that agent's page. A client already commissioned is never shown (no
hint of paying twice). The unit credit can bump the agent's tier (and therefore
their commission on their OTHER real cleared debt that month), but the EPF
client's own debt is never added to total_cleared_debt, so it adds no commission
dollars on its own."""

import io
from types import SimpleNamespace

import openpyxl
import pytest

from app.calculator import calculate_agent_commission
from app.cordoba_parser import parse_cordoba_payout
from app.models import CommissionPeriod, AgentCommission, ClientRecord, EpfClient
from app.routes import _apply_epf_rows

FAKE_FILE = SimpleNamespace(filename="cordoba_payouts.xlsx")


def build_cordoba_xlsx(epf_rows):
    """Minimal workbook with the three expected tabs; EPF rows are (contact_id, name, cleared_date)."""
    wb = openpyxl.Workbook()
    first = wb.active
    first.title = "First Pays"
    first.append(["ID", "Full Name"])
    epf = wb.create_sheet("EPF")
    epf.append(["Contact ID", "Full Name", "Cleared Date"])
    for row in epf_rows:
        epf.append(list(row))
    cb = wb.create_sheet("Chargebacks")
    cb.append(["ID", "Full Name", "Dropped Date"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def seed_client(db, crm_id="7001", agent="Maria", is_cleared=False):
    period = CommissionPeriod(period_label="2026-05", filename="crm.csv", total_agents=1)
    db.session.add(period)
    db.session.flush()
    agent_row = AgentCommission(
        period_id=period.id, agent_name=agent,
        units_cleared=1, total_cleared_debt=10_000.0, cancellation_rate=0.0,
        hourly_draw=0.0, raw_tier=1, adjusted_tier=1, tier_rate=0.01,
        gross_commission=100.0, clawback_amount=0.0, net_commission=100.0,
        payout=100.0, payout_type="commission", source="crm", notes="",
    )
    db.session.add(agent_row)
    db.session.flush()
    db.session.add(ClientRecord(
        period_id=period.id, agent_commission_id=agent_row.id,
        crm_id=crm_id, agent_name=agent, client_name="Jane Roe",
        enrolled_debt=10_000.0, is_cleared=is_cleared, is_pending=not is_cleared,
        status="Pending Affiliate Cancellation" if not is_cleared else "Active",
    ))
    db.session.commit()
    return agent_row


def seed_agent_period(db, period_label, agent, units_cleared, total_cleared_debt):
    """Seed a CommissionPeriod + AgentCommission with real (non-EPF) cleared units,
    using the same tier math as production so tests exercise a realistic tier boundary."""
    period = CommissionPeriod.query.filter_by(period_label=period_label).first()
    if not period:
        period = CommissionPeriod(period_label=period_label, filename="crm.csv", total_agents=0)
        db.session.add(period)
        db.session.flush()
    result = calculate_agent_commission(
        agent_name=agent, units_cleared=units_cleared,
        total_cleared_debt=total_cleared_debt, cancellation_rate_pct=0.0,
    )
    agent_row = AgentCommission(
        period_id=period.id, agent_name=agent,
        units_cleared=units_cleared, total_cleared_debt=total_cleared_debt,
        cancellation_rate=0.0, hourly_draw=0.0,
        raw_tier=result["raw_tier"], adjusted_tier=result["adjusted_tier"],
        tier_rate=result["tier_rate"], gross_commission=result["gross_commission"],
        clawback_amount=0.0, net_commission=result["gross_commission"],
        payout=result["payout"], payout_type=result["payout_type"],
        source="crm", notes=result["notes"],
    )
    db.session.add(agent_row)
    period.total_agents = (period.total_agents or 0) + 1
    db.session.commit()
    return agent_row


def test_parser_extracts_epf_rows_with_cleared_date():
    data = build_cordoba_xlsx([("7001", "Jane Roe", "06/15/2026")])
    parsed = parse_cordoba_payout(data)
    assert parsed["errors"] == []
    assert parsed["epf_rows"] == [
        {"crm_id": "7001", "client_name": "Jane Roe", "cleared_date": "06/15/2026"},
    ]
    # EPF still confirms Cordoba paid us (feeds the CordobaPaidClient ledger)
    assert {"crm_id": "7001", "client_name": "Jane Roe", "source": "epf"} in parsed["paid_ids"]


def test_epf_row_stored_for_matched_agent_and_month(db):
    seed_client(db, is_cleared=False)
    added, skipped, unmatched, missing = _apply_epf_rows(
        FAKE_FILE, {"epf_rows": [{"crm_id": "7001", "client_name": "Jane Roe",
                                  "cleared_date": "06/15/2026"}]})
    db.session.commit()
    assert (added, skipped, unmatched, missing) == (1, 0, 0, 0)
    entry = EpfClient.query.one()
    assert entry.agent_name == "Maria"
    assert entry.period_label == "2026-06"      # from the EPF Cleared Date
    assert entry.cleared_date == "06/15/2026"


def test_epf_never_changes_commission_when_period_not_created_yet(db):
    # ClientRecord used only to resolve the agent name lives in a different month
    # (2026-05) than the EPF row's own Cleared Date (2026-06). No CommissionPeriod
    # exists yet for 2026-06, so there's nothing to recompute — the unit is picked up
    # later once a CRM upload actually creates that period (see crm_parser tests).
    seed_client(db, is_cleared=False)
    _apply_epf_rows(FAKE_FILE, {"epf_rows": [{"crm_id": "7001", "client_name": "Jane Roe",
                                              "cleared_date": "06/15/2026"}]})
    db.session.commit()
    assert CommissionPeriod.query.filter_by(period_label="2026-06").first() is None
    # 2026-05 period (unrelated month) is untouched
    assert AgentCommission.query.filter_by(period_id=CommissionPeriod.query.filter_by(
        period_label="2026-05").first().id).count() == 1


def test_epf_adds_unit_toward_tier_without_adding_debt(db):
    # Agent already has 20 real cleared units (Tier 1 upper bound) in 2026-06 —
    # the same month the EPF row's Cleared Date will land in.
    agent_row = seed_agent_period(db, "2026-06", "Maria", units_cleared=20, total_cleared_debt=100_000.0)
    assert agent_row.tier_rate == 0.01
    assert agent_row.gross_commission == 1000.0

    # ClientRecord (any period) so _apply_epf_rows can resolve crm_id 7001 -> Maria
    seed_client(db, crm_id="9999", agent="Maria", is_cleared=False)  # unrelated client
    period_05 = CommissionPeriod.query.filter_by(period_label="2026-05").first()
    db.session.add(ClientRecord(
        period_id=period_05.id, crm_id="7001", agent_name="Maria",
        client_name="Jane Roe", enrolled_debt=5_000.0, is_cleared=False, is_pending=True,
    ))
    db.session.commit()

    added, *_ = _apply_epf_rows(FAKE_FILE, {"epf_rows": [
        {"crm_id": "7001", "client_name": "Jane Roe", "cleared_date": "06/15/2026"},
    ]})
    db.session.commit()
    assert added == 1

    refreshed = db.session.get(AgentCommission, agent_row.id)
    # Tier bumped from 1 to 2 (21 units), rate 1.25%, but total_cleared_debt is
    # unchanged — the EPF client's own $5,000 debt was never added.
    assert refreshed.units_cleared == 21
    assert refreshed.epf_units == 1
    assert refreshed.total_cleared_debt == 100_000.0
    assert refreshed.raw_tier == 2
    assert refreshed.tier_rate == 0.0125
    assert refreshed.gross_commission == 1_250.0
    assert refreshed.net_commission == 1_250.0
    assert "EPF: +1 unit(s) credited toward tier" in refreshed.notes


def test_epf_recompute_is_idempotent_across_multiple_uploads(db):
    agent_row = seed_agent_period(db, "2026-06", "Maria", units_cleared=20, total_cleared_debt=100_000.0)
    period_05 = CommissionPeriod.query.filter_by(period_label="2026-05").first()
    if not period_05:
        period_05 = CommissionPeriod(period_label="2026-05", filename="crm.csv", total_agents=0)
        db.session.add(period_05)
        db.session.flush()
    for crm_id in ("7001", "7002"):
        db.session.add(ClientRecord(
            period_id=period_05.id, crm_id=crm_id, agent_name="Maria",
            client_name="Client " + crm_id, enrolled_debt=5_000.0,
            is_cleared=False, is_pending=True,
        ))
    db.session.commit()

    # First EPF file: one row
    _apply_epf_rows(FAKE_FILE, {"epf_rows": [
        {"crm_id": "7001", "client_name": "Client 7001", "cleared_date": "06/15/2026"},
    ]})
    db.session.commit()
    mid = db.session.get(AgentCommission, agent_row.id)
    assert mid.units_cleared == 21
    assert mid.epf_units == 1

    # Second EPF file: a new row for the same agent/month
    _apply_epf_rows(FAKE_FILE, {"epf_rows": [
        {"crm_id": "7002", "client_name": "Client 7002", "cleared_date": "06/20/2026"},
    ]})
    db.session.commit()
    refreshed = db.session.get(AgentCommission, agent_row.id)
    assert refreshed.units_cleared == 22   # base 20 + 2 EPF, not 23
    assert refreshed.epf_units == 2
    assert refreshed.total_cleared_debt == 100_000.0
    assert refreshed.notes.count("EPF:") == 1  # note segment replaced, not duplicated

    # Re-uploading the same file again is a no-op and doesn't change the count
    _apply_epf_rows(FAKE_FILE, {"epf_rows": [
        {"crm_id": "7002", "client_name": "Client 7002", "cleared_date": "06/20/2026"},
    ]})
    db.session.commit()
    refreshed = db.session.get(AgentCommission, agent_row.id)
    assert refreshed.epf_units == 2


def test_already_commissioned_client_is_skipped(db):
    seed_client(db, is_cleared=True)
    added, skipped, unmatched, missing = _apply_epf_rows(
        FAKE_FILE, {"epf_rows": [{"crm_id": "7001", "client_name": "Jane Roe",
                                  "cleared_date": "06/15/2026"}]})
    assert (added, skipped, unmatched, missing) == (0, 1, 0, 0)
    assert EpfClient.query.count() == 0


def test_unknown_contact_id_is_skipped(db):
    seed_client(db)
    added, skipped, unmatched, missing = _apply_epf_rows(
        FAKE_FILE, {"epf_rows": [{"crm_id": "9999", "client_name": "Nobody",
                                  "cleared_date": "06/15/2026"}]})
    assert (added, skipped, unmatched, missing) == (0, 0, 1, 0)


def test_missing_cleared_date_is_skipped(db):
    seed_client(db)
    added, skipped, unmatched, missing = _apply_epf_rows(
        FAKE_FILE, {"epf_rows": [{"crm_id": "7001", "client_name": "Jane Roe",
                                  "cleared_date": None}]})
    assert (added, skipped, unmatched, missing) == (0, 0, 0, 1)


def test_reupload_is_a_noop(db):
    seed_client(db)
    row = {"epf_rows": [{"crm_id": "7001", "client_name": "Jane Roe",
                         "cleared_date": "06/15/2026"}]}
    _apply_epf_rows(FAKE_FILE, row)
    db.session.commit()
    added, *_ = _apply_epf_rows(FAKE_FILE, row)
    assert added == 0
    assert EpfClient.query.count() == 1
