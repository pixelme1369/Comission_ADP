import csv
import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from flask import Blueprint, render_template, request, redirect, url_for, flash, Response
from app import db
from app.models import (
    CommissionPeriod, AgentCommission, ClientRecord, CordobaPaidClient,
    CordobaChargedBackClient, CordobaChargebackMatchedClient, CordobaChargebackEntry,
)
from commission_core.crm_parser import parse_crm_and_calculate, _parse_date, _period_of
from commission_core.cordoba_parser import parse_cordoba_payout
from commission_core.commission_history_parser import parse_commission_history
from commission_core.calculator import (
    calculate_clawback_amount, get_fixed_rate, units_to_next_tier, commission_gain_at_next_tier,
)

bp = Blueprint("main", __name__)

ALLOWED_EXTENSIONS = {"csv"}
ALLOWED_XLSX_EXTENSIONS = {"xlsx"}
ALLOWED_HISTORY_EXTENSIONS = {"xlsx", "csv"}


def _allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def _units_to_next_tier_map(agents):
    return {a.id: units_to_next_tier(a.units_cleared, a.agent_name) for a in agents}


CRM_SOURCES = ("crm", "manual")
HISTORY_SOURCES = ("history_import",)


def _group_periods_by_filename(periods):
    """Groups CommissionPeriod rows by their source filename (a multi-month CRM
    export or history-ledger upload can create several periods from one file),
    for the index page's per-upload-type "recent uploads" lists — newest
    upload first."""
    groups = {}
    for p in periods:
        groups.setdefault(p.filename or "(unknown)", []).append(p)
    return sorted(
        ({"filename": name, "periods": sorted(rows, key=lambda r: r.period_label)}
         for name, rows in groups.items()),
        key=lambda g: max(r.uploaded_at for r in g["periods"]), reverse=True,
    )


def _delete_periods_by_filename(filename, source_values):
    """Deletes every CommissionPeriod with this filename whose agents were
    created by one of the given upload types (source_values) — lets the index
    page delete/reset an entire multi-month upload in one action instead of
    one period at a time. Cascades to AgentCommission/ClientRecord via the
    model relationships, same as deleting a period individually. Returns the
    list of period_labels actually deleted."""
    periods = CommissionPeriod.query.filter_by(filename=filename).all()
    deleted_labels = []
    for period in periods:
        if period.agents and period.agents[0].source in source_values:
            deleted_labels.append(period.period_label)
            db.session.delete(period)
    db.session.commit()
    return deleted_labels


def _allowed_xlsx_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_XLSX_EXTENSIONS


def _allowed_history_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_HISTORY_EXTENSIONS


def _new_client_record(period_id, agent_commission_id, cr, **overrides):
    """Build a ClientRecord from a parser client dict. Every upload flow saves clients
    through this one helper so the field mapping can never drift between flows;
    flow-specific values (clawback fields, cordoba_paid, ...) come in as overrides."""
    fields = dict(
        period_id=period_id,
        agent_commission_id=agent_commission_id,
        crm_id=cr.get("crm_id"),
        agent_name=cr["agent_name"],
        client_name=cr.get("client_name"),
        email=cr.get("email"),
        phone=cr.get("phone"),
        stage=cr.get("stage"),
        status=cr.get("status"),
        submitted_date=cr.get("submitted_date"),
        enrolled_date=cr.get("enrolled_date"),
        first_payment_date=cr.get("first_payment_date"),
        first_payment_cleared_date=cr.get("first_payment_cleared_date"),
        second_payment_cleared_date=cr.get("second_payment_cleared_date"),
        dropped_date=cr.get("dropped_date"),
        pay_freq=cr.get("pay_freq"),
        payments_made=cr.get("payments_made", 0),
        nsf_count=cr.get("nsf_count", 0),
        enrolled_debt=cr.get("enrolled_debt", 0.0),
        credit_score=cr.get("credit_score"),
        is_low_credit=cr.get("is_low_credit", False),
        is_cleared=cr.get("is_cleared", False),
        is_pending=cr.get("is_pending", False),
        is_cancelled=cr.get("is_cancelled", False),
        commission_on_client=cr.get("commission_on_client", 0.0),
    )
    fields.update(overrides)
    return ClientRecord(**fields)


@bp.route("/")
def index():
    recent_periods = CommissionPeriod.query.order_by(CommissionPeriod.uploaded_at.desc()).limit(12).all()
    all_periods = CommissionPeriod.query.all()
    crm_periods = [p for p in all_periods if p.agents and p.agents[0].source in CRM_SOURCES]
    history_periods = [p for p in all_periods if p.agents and p.agents[0].source in HISTORY_SOURCES]
    return render_template(
        "index.html", periods=recent_periods,
        crm_uploads=_group_periods_by_filename(crm_periods),
        history_uploads=_group_periods_by_filename(history_periods),
        cordoba_uploads=_list_cordoba_uploads(),
    )


@bp.route("/reset-all", methods=["POST"])
def reset_all():
    # Per-object delete (not a bulk Query.delete()) so the ORM cascade on
    # CommissionPeriod actually fires and takes AgentCommission + ClientRecord with it.
    for period in CommissionPeriod.query.all():
        db.session.delete(period)
    CordobaPaidClient.query.delete()
    CordobaChargedBackClient.query.delete()
    CordobaChargebackMatchedClient.query.delete()
    db.session.commit()
    flash("All commission data has been reset.", "success")
    return redirect(url_for("main.index"))


@bp.route("/upload-crm", methods=["POST"])
def upload_crm():
    file = request.files.get("csv_file")
    if not file or file.filename == "":
        flash("No file selected.", "error")
        return redirect(url_for("main.index"))
    if not _allowed_file(file.filename):
        flash("Only .csv files are accepted.", "error")
        return redirect(url_for("main.index"))

    file_bytes = file.read()

    # Collect crm_ids already saved as cleared so the parser can detect late activations
    already_cleared_crm_ids = {
        r[0] for r in db.session.query(ClientRecord.crm_id)
        .filter(ClientRecord.is_cleared.is_(True)) if r[0]
    }
    # Collect crm_ids Cordoba has already confirmed paying (from a prior payout upload)
    already_cordoba_paid_ids = {r[0] for r in db.session.query(CordobaPaidClient.crm_id)}
    # Collect crm_ids already clawed back — from a Cordoba Chargebacks-tab upload OR a
    # PRIOR CRM upload's own clawback detection. Both matter now: clawbacks land in the
    # client's own Dropped Date month (owner policy, confirmed August 2026), which is
    # stable/unchanging across every future upload — without this second source, a
    # full-history re-upload would re-detect and re-apply the SAME clawback every time,
    # since the target period never moves forward the way "latest period in file" used to.
    already_charged_back_crm_ids = {
        r[0] for r in db.session.query(CordobaChargedBackClient.crm_id) if r[0]
    } | {
        r[0] for r in db.session.query(ClientRecord.crm_id)
        .filter(ClientRecord.clawback_applied.is_(True)) if r[0]
    }
    # crm_ids already saved as low-credit (Credit Score <= 500) cleared clients, so a
    # later drop on that same client never triggers a clawback — they were never paid
    # any commission to begin with.
    already_low_credit_crm_ids = {
        r[0] for r in db.session.query(ClientRecord.crm_id)
        .filter(ClientRecord.is_low_credit.is_(True)) if r[0]
    }

    period_results = parse_crm_and_calculate(
        file_bytes, file.filename, already_cleared_crm_ids, already_charged_back_crm_ids,
        already_low_credit_crm_ids,
    )

    saved_period_ids = []
    updated_period_ids = []
    shown_errors = set()

    for parsed in period_results:
        # Show row-level warnings once (they repeat across periods)
        for err in parsed.get("errors", []):
            if err not in shown_errors:
                flash(err, "error")
                shown_errors.add(err)

        if not parsed["results"] or not parsed["period_label"]:
            continue

        period_label = parsed["period_label"]
        existing = CommissionPeriod.query.filter_by(period_label=period_label).first()

        if not existing:
            period = CommissionPeriod(
                period_label=period_label,
                filename=file.filename,
                total_agents=len(parsed["results"]),
            )
            db.session.add(period)
            db.session.flush()

            # Save agent commission records
            # Strip internal keys before saving to model
            agent_obj_map = {}  # agent_name → AgentCommission
            for r in parsed["results"]:
                cleared_clients = r.pop("_cleared_clients", [])
                all_period_clients = r.pop("_all_period_clients", [])
                clawback_clients = r.pop("_clawback_clients", [])
                r.pop("_period_label", None)

                agent_obj = AgentCommission(period_id=period.id, **r)
                db.session.add(agent_obj)
                db.session.flush()
                agent_obj_map[r["agent_name"]] = {
                    "obj": agent_obj,
                    "cleared_clients": cleared_clients,
                    "all_period_clients": all_period_clients,
                    "clawback_clients": clawback_clients,
                }

            # Save individual client records
            for agent_name, data in agent_obj_map.items():
                agent_obj = data["obj"]

                # Clients that belong to this period (cleared, pending, same-month cancel)
                for cr in data["all_period_clients"]:
                    db.session.add(_new_client_record(
                        period.id, agent_obj.id, cr,
                        is_late_activation=cr.get("is_late_activation", False),
                        original_cleared_period=cr.get("original_cleared_period"),
                        cordoba_paid=cr.get("crm_id") in already_cordoba_paid_ids,
                    ))

                # Clawback clients — these cleared in a prior month, cancelled this month
                for cr in data["clawback_clients"]:
                    db.session.add(_new_client_record(
                        period.id, agent_obj.id, cr,
                        is_cleared=False,
                        is_pending=False,
                        is_cancelled=True,
                        commission_on_client=0.0,
                        clawback_applied=True,
                        clawback_period_id=period.id,
                        clawback_amount=cr.get("clawback_amount", 0.0),
                    ))

            saved_period_ids.append((period.id, period_label, len(parsed["results"])))
            continue

        # Period already exists. Genuine new cleared/safe-cancel activity for that
        # month is NOT re-imported here (unchanged protection against double-counting
        # an already-recorded month — delete it first to re-import). But any NEW
        # clawback found for this month IS still applied: now that clawbacks target
        # the client's own Dropped Date month (owner policy, confirmed August 2026),
        # that month is very often one that already exists on file — that's no longer
        # a reason to silently lose the clawback. Applied via find-or-create, mirroring
        # how the separate Cordoba-chargeback flow has always attached a deduction to
        # an existing period (see _get_or_create_agent_period_row below).
        period_had_new_units = any(r.get("units_cleared", 0) > 0 for r in parsed["results"])
        new_clawback_total = 0.0
        new_clawback_count = 0

        for r in parsed["results"]:
            clawback_clients = r.get("_clawback_clients", [])
            if not clawback_clients:
                continue

            agent_name = r["agent_name"]
            cb_ids = {c["crm_id"] for c in clawback_clients if c.get("crm_id")}
            already_recorded = {
                x[0] for x in db.session.query(ClientRecord.crm_id).filter(
                    ClientRecord.crm_id.in_(cb_ids),
                    ClientRecord.clawback_applied.is_(True),
                )
            } if cb_ids else set()
            new_cb = [c for c in clawback_clients
                      if not c.get("crm_id") or c["crm_id"] not in already_recorded]
            if not new_cb:
                continue

            agent_row = AgentCommission.query.filter_by(
                period_id=existing.id, agent_name=agent_name).first()
            if not agent_row:
                agent_row = AgentCommission(
                    period_id=existing.id, agent_name=agent_name,
                    units_cleared=0, total_cleared_debt=0.0, cancellation_rate=0.0, hourly_draw=0.0,
                    raw_tier=0, adjusted_tier=0, tier_rate=0.0, gross_commission=0.0,
                    clawback_amount=0.0, net_commission=0.0, payout=0.0, payout_type="none",
                    quality_bonus_eligible=False, cancellation_penalty_applied=False, nsf_flagged=False,
                    pending_units=0, pending_debt=0.0, source="crm", notes="",
                )
                db.session.add(agent_row)
                db.session.flush()
                existing.total_agents = (existing.total_agents or 0) + 1

            total_cb = round(sum(c.get("clawback_amount", 0.0) for c in new_cb), 2)
            agent_row.clawback_amount = round((agent_row.clawback_amount or 0.0) + total_cb, 2)
            agent_row.net_commission = max(0.0, round(agent_row.gross_commission - agent_row.clawback_amount, 2))
            agent_row.notes = (agent_row.notes or "") + \
                f" | Clawback -${total_cb:,.2f} from {len(new_cb)} previously-paid cancelled client(s)"

            for cr in new_cb:
                db.session.add(_new_client_record(
                    existing.id, agent_row.id, cr,
                    is_cleared=False, is_pending=False, is_cancelled=True,
                    commission_on_client=0.0, clawback_applied=True,
                    clawback_period_id=existing.id, clawback_amount=cr.get("clawback_amount", 0.0),
                ))
            new_clawback_total += total_cb
            new_clawback_count += len(new_cb)

        if period_had_new_units:
            flash(
                f"Period {period_label} already exists (uploaded {existing.uploaded_at.strftime('%Y-%m-%d')}). "
                "New cleared/safe-cancel activity for that month was NOT re-imported — delete it first "
                "if you need to re-import that month's calculated commissions.", "error",
            )
        if new_clawback_count:
            updated_period_ids.append(existing.id)
            flash(
                f"Period {period_label} already existed — applied {new_clawback_count} new "
                f"clawback(s) totaling ${new_clawback_total:,.2f} to it.", "success",
            )

    # One commit for the whole file: either every new period saves, or none do.
    # (Per-period commits could leave a half-imported file if a later period failed.)
    db.session.commit()

    if not saved_period_ids and not updated_period_ids:
        return redirect(url_for("main.index"))

    for pid, plabel, count in saved_period_ids:
        flash(f"CRM import: {count} agents processed for period {plabel}.", "success")

    all_period_ids = [pid for pid, _, _ in saved_period_ids] + updated_period_ids
    if len(all_period_ids) == 1:
        return redirect(url_for("main.period_detail", period_id=all_period_ids[0]))
    return redirect(url_for("main.history"))


def _apply_cordoba_paid_flags(file, parsed):
    """
    Check OUR existing ClientRecord IDs against Cordoba's First Pays / EPF tabs (not the
    reverse) — for every client we already have on file whose ID shows up in either tab,
    flip cordoba_paid = True.
    """
    incoming_ids = {row["crm_id"] for row in parsed["paid_ids"] if row["crm_id"]}
    if not incoming_ids:
        return 0, 0

    already_known_ids = {
        p.crm_id for p in CordobaPaidClient.query.filter(CordobaPaidClient.crm_id.in_(incoming_ids)).all()
    }

    new_count = 0
    seen_this_file = set()
    for row in parsed["paid_ids"]:
        crm_id = row["crm_id"]
        if not crm_id or crm_id in already_known_ids or crm_id in seen_this_file:
            continue
        seen_this_file.add(crm_id)
        db.session.add(CordobaPaidClient(
            crm_id=crm_id, client_name=row.get("client_name"), source=row["source"],
            uploaded_filename=file.filename,
        ))
        new_count += 1

    # Deliberately not filtering on the current cordoba_paid value here (e.g. "IS False").
    # A row can end up with NULL instead of False if it was ever inserted while this
    # column didn't exist in the model (this happened once, see CLAUDE.md) — "IS False"
    # would silently skip NULL rows forever since NULL isn't equal to False in SQL.
    # Just unconditionally set every matching ID to True; re-setting an already-True row
    # is harmless.
    flipped = ClientRecord.query.filter(ClientRecord.crm_id.in_(incoming_ids)).update(
        {"cordoba_paid": True}, synchronize_session=False
    )

    return new_count, flipped


def _client_label(client_name, crm_id):
    """'Eddie Ramos (ID 1181065497)' — skip messages must carry the ID so a client
    can be cross-checked against the CRM export directly."""
    return f"{client_name} (ID {crm_id})" if client_name else crm_id


def _get_or_create_agent_period_row(period_label, agent_name, filename):
    """Find (or create a zero-unit) AgentCommission row to carry a clawback that has
    no cleared units of its own in this period — mirrors the CRM-clawback holding entry.
    Only called by the Cordoba chargeback flow (source="cordoba", not "crm"), so the
    index page's per-upload-type "recent uploads" lists don't mistake a bare Cordoba
    holding period for a CRM export upload."""
    period = CommissionPeriod.query.filter_by(period_label=period_label).first()
    if not period:
        period = CommissionPeriod(period_label=period_label, filename=filename, total_agents=0)
        db.session.add(period)
        db.session.flush()

    agent_row = AgentCommission.query.filter_by(period_id=period.id, agent_name=agent_name).first()
    if not agent_row:
        agent_row = AgentCommission(
            period_id=period.id, agent_name=agent_name,
            units_cleared=0, total_cleared_debt=0.0, cancellation_rate=0.0, hourly_draw=0.0,
            raw_tier=0, adjusted_tier=0, tier_rate=0.0, gross_commission=0.0,
            clawback_amount=0.0, net_commission=0.0, payout=0.0, payout_type="none",
            quality_bonus_eligible=False, cancellation_penalty_applied=False, nsf_flagged=False,
            pending_units=0, pending_debt=0.0, source="cordoba", notes="",
        )
        db.session.add(agent_row)
        db.session.flush()
        period.total_agents = (period.total_agents or 0) + 1

    return period, agent_row


def _mark_cordoba_chargeback_matches(file, parsed):
    """
    Display-only companion to _apply_cordoba_chargebacks: for every ID in the
    Chargebacks tab, check it against ALL of our own commission reports
    (ClientRecord.crm_id, any period, any status — not gated on is_cleared, confirmed
    payout, or having a Dropped Date on file). Any match is recorded in
    CordobaChargebackMatchedClient forever, which drives the per-client "Cordoba
    Clawback" Yes/No badge next to "Cordoba Payout" on the agent detail page.
    This intentionally does NOT gate on anything _apply_cordoba_chargebacks checks —
    the badge should show Yes as soon as we recognize the client, even if the actual
    dollar deduction is still blocked (most commonly: no Dropped Date on file yet).
    Returns (newly_marked_count, unmatched_labels).
    """
    chargebacks = parsed.get("chargebacks", [])
    incoming_ids = {row["crm_id"] for row in chargebacks if row["crm_id"]}
    if not incoming_ids:
        return 0, []

    already_marked = {
        r[0] for r in db.session.query(CordobaChargebackMatchedClient.crm_id)
        .filter(CordobaChargebackMatchedClient.crm_id.in_(incoming_ids))
    }
    known_ids = {
        r[0] for r in db.session.query(ClientRecord.crm_id)
        .filter(ClientRecord.crm_id.in_(incoming_ids))
    }

    marked = 0
    unmatched = []
    seen_this_file = set()
    for row in chargebacks:
        crm_id = row["crm_id"]
        if not crm_id or crm_id in seen_this_file:
            continue
        seen_this_file.add(crm_id)

        if crm_id not in known_ids:
            unmatched.append(_client_label(row.get("client_name"), crm_id))
            continue
        if crm_id in already_marked:
            continue

        db.session.add(CordobaChargebackMatchedClient(
            crm_id=crm_id, client_name=row.get("client_name"), uploaded_filename=file.filename,
        ))
        marked += 1

    return marked, unmatched


def _apply_cordoba_chargebacks(file, parsed):
    """
    Cross-reference Cordoba's Chargebacks tab against OUR OWN ClientRecord history — the
    tab has no agent/rep column, so "who gets charged back" comes from looking up each
    charged-back client ID in our records, not from the file. Two things must both be
    true before we claw anything back:
      1. We ever paid an agent commission on that client (ClientRecord.is_cleared was True)
      2. Cordoba has confirmed paying US on it at some point — the crm_id has appeared in
         a First Pays or EPF tab, ever (CordobaPaidClient ledger, not limited to this file)
    A chargeback logically can't exist without a prior payment, so #2 mainly catches data
    gaps (Cordoba's chargeback tab referencing an ID whose original payout confirmation we
    never uploaded) rather than filtering out real clawbacks.
    Claw back that agent's commission in the month the client dropped — unconditionally,
    regardless of the safe-payment-threshold that protects agents in the CRM-driven
    clawback flow, since Cordoba taking the marketing payout back from us is independent
    of that policy. Each crm_id is recorded in CordobaChargedBackClient forever so
    re-uploading this file, or a later CRM upload reflecting the same drop, never claws
    the agent back twice.
    """
    chargebacks = parsed.get("chargebacks", [])
    incoming_ids = {row["crm_id"] for row in chargebacks if row["crm_id"]}
    if not incoming_ids:
        return 0, 0.0, [], [], [], []

    already_charged_back = {
        c.crm_id for c in
        CordobaChargedBackClient.query.filter(CordobaChargedBackClient.crm_id.in_(incoming_ids)).all()
    }
    confirmed_paid_ids = {
        p.crm_id for p in
        CordobaPaidClient.query.filter(CordobaPaidClient.crm_id.in_(incoming_ids)).all()
    }
    # Third gate — never claw back a client who was ALREADY clawed back through any
    # other path: a CRM upload that reflected the drop (clawback_applied ClientRecord),
    # or a prior manager's "To subtract" row from a commission-history import (which
    # also creates a clawback_applied ClientRecord). Without this, the Cordoba
    # Chargebacks tab arriving AFTER the CRM export already clawed the agent back
    # would deduct the same client a second time — the CordobaChargedBackClient
    # ledger above only guards the Cordoba-first ordering.
    already_clawed_elsewhere = {
        r[0] for r in
        db.session.query(ClientRecord.crm_id).filter(
            ClientRecord.crm_id.in_(incoming_ids),
            ClientRecord.clawback_applied.is_(True),
        )
    }

    applied_count = 0
    total_clawed_back = 0.0
    skipped_not_commissioned = []
    skipped_not_confirmed_paid = []
    skipped_already_clawed = []
    skipped_no_dropped_date = []
    seen_this_file = set()

    for row in chargebacks:
        crm_id = row["crm_id"]
        if not crm_id or crm_id in already_charged_back or crm_id in seen_this_file:
            continue
        seen_this_file.add(crm_id)

        if crm_id in already_clawed_elsewhere:
            # Agent was already deducted for this client (CRM upload or history import).
            skipped_already_clawed.append(_client_label(row.get("client_name"), crm_id))
            continue

        client_rec = (
            ClientRecord.query.filter_by(crm_id=crm_id, is_cleared=True)
            .order_by(ClientRecord.id.desc()).first()
        )
        if not client_rec:
            # We never recorded this client as cleared/commissioned — nothing to claw back.
            skipped_not_commissioned.append(_client_label(row.get("client_name"), crm_id))
            continue

        if crm_id not in confirmed_paid_ids:
            # Cordoba's own First Pays/EPF tabs never confirmed paying us on this ID —
            # don't claw back an agent for a payout we can't verify we received.
            skipped_not_confirmed_paid.append(_client_label(row.get("client_name"), crm_id))
            continue

        # Owner policy (confirmed July 2026): the Chargebacks tab's own Dropped Date
        # column is not used — only OUR OWN CRM-recorded dropped_date places the
        # deduction. If we don't have one yet (the drop hasn't shown up in a CRM
        # upload), there's nowhere to put the clawback yet — skip until it does.
        dropped_dt = _parse_date(client_rec.dropped_date or "")
        dropped_period = _period_of(dropped_dt)
        if not dropped_period:
            skipped_no_dropped_date.append(_client_label(row.get("client_name"), crm_id))
            continue

        agent_name = client_rec.agent_name
        # Owner policy: the Chargebacks tab's Marketing Payout Debt column is not used —
        # client debt comes only from our own CRM-recorded Enrolled Debt.
        client_debt = client_rec.enrolled_debt or 0.0
        orig_agent_row = client_rec.agent_commission

        if orig_agent_row and orig_agent_row.units_cleared > 0:
            cb = calculate_clawback_amount(
                orig_agent_row.units_cleared,
                orig_agent_row.total_cleared_debt,
                orig_agent_row.gross_commission,
                orig_agent_row.cancellation_rate,
                client_debt,
                agent_name=agent_name,
            )
        else:
            cb = round(client_debt * (get_fixed_rate(agent_name) or 0.01), 2)

        if cb <= 0:
            continue

        period, agent_row = _get_or_create_agent_period_row(dropped_period, agent_name, file.filename)

        agent_row.clawback_amount = round((agent_row.clawback_amount or 0.0) + cb, 2)
        agent_row.net_commission = max(0.0, round(agent_row.gross_commission - agent_row.clawback_amount, 2))
        note = f"Cordoba chargeback: -${cb:,.2f} for {client_rec.client_name or crm_id} (ID {crm_id})"
        agent_row.notes = f"{agent_row.notes} | {note}" if agent_row.notes else note

        db.session.add(ClientRecord(
            period_id=period.id,
            agent_commission_id=agent_row.id,
            crm_id=crm_id,
            agent_name=agent_name,
            client_name=client_rec.client_name,
            email=client_rec.email,
            phone=client_rec.phone,
            stage=client_rec.stage,
            status="Cordoba Chargeback",
            enrolled_date=client_rec.enrolled_date,
            first_payment_cleared_date=client_rec.first_payment_cleared_date,
            dropped_date=client_rec.dropped_date,
            pay_freq=client_rec.pay_freq,
            payments_made=client_rec.payments_made,
            enrolled_debt=client_debt,
            is_cleared=False,
            is_pending=False,
            is_cancelled=True,
            commission_on_client=0.0,
            clawback_applied=True,
            clawback_period_id=period.id,
            clawback_amount=cb,
        ))

        db.session.add(CordobaChargedBackClient(
            crm_id=crm_id, client_name=client_rec.client_name, agent_name=agent_name,
            clawback_amount=cb, dropped_period=dropped_period, uploaded_filename=file.filename,
        ))

        applied_count += 1
        total_clawed_back += cb

    return (applied_count, round(total_clawed_back, 2),
            skipped_not_commissioned, skipped_not_confirmed_paid, skipped_already_clawed,
            skipped_no_dropped_date)


def _list_cordoba_chargebacks(file, parsed):
    """
    Display-only companion to _apply_cordoba_chargebacks (owner request, July 2026):
    for every ID in the Chargebacks tab, look up the agent and dropped month from OUR
    OWN ClientRecord history and record a verbatim snapshot of that file row (Assigned
    Company, dates, Pay Freq., Payments Made, Marketing Payout Debt, Marketing Payment
    Cleared/Chargeback, the file's own Dropped Date) as a CordobaChargebackEntry, shown/
    exported as "Cordoba Charge back". This never touches gross_commission,
    net_commission, or clawback_amount — it's purely informational, for the row to be
    reconciled by hand.

    Gated on the client having actually been paid at some point (some ClientRecord row
    for this crm_id has is_cleared=True or clawback_applied=True) — a client who
    dropped before their commission was ever paid (same_month_cancel: "Cancelled — Not
    Paid") has nothing to charge back, so listing them here would misrepresent a
    Cordoba chargeback as reconciliation-worthy when no commission ever went out.
    If the ID doesn't match any ClientRecord we have, we don't have a dropped date on
    file for it yet (needed to know which period to show it under), or the client was
    never actually paid, it's skipped and reported back to the uploader so it isn't
    silently dropped.

    UPSERT, not insert-once (owner request, confirmed): the first version of this only
    ever inserted a crm_id's entry once — re-uploading a LATER Cordoba payout file with
    updated figures for a client already listed (more Payments Made, a corrected
    Marketing Payout Debt, a later Marketing Payment Chargeback date, etc.) was
    silently a no-op, leaving the reconciliation table showing stale numbers forever.
    An existing CordobaChargebackEntry for a crm_id is now overwritten in place with
    this file's row instead, including uploaded_filename — so the entry always reflects
    whichever Cordoba file most recently reported that client, and deleting an OLDER
    upload no longer touches an entry a newer upload has since refreshed (see
    _delete_cordoba_upload). Safe to do unconditionally: this table is purely
    informational (never gross/net/clawback), unlike CordobaChargedBackClient's
    real-money ledger, which deliberately still guards against ever re-applying the
    same deduction twice — that policy is untouched here.
    Returns (newly_listed_count, updated_count, total_marketing_payout_debt, skipped_no_match_labels).
    """
    chargebacks = parsed.get("chargebacks", [])
    incoming_ids = {row["crm_id"] for row in chargebacks if row["crm_id"]}
    if not incoming_ids:
        return 0, 0, 0.0, []

    existing_by_id = {
        e.crm_id: e for e in
        CordobaChargebackEntry.query.filter(CordobaChargebackEntry.crm_id.in_(incoming_ids))
    }

    listed = 0
    updated = 0
    total = 0.0
    skipped_no_match = []
    seen_this_file = set()
    for row in chargebacks:
        crm_id = row["crm_id"]
        if not crm_id or crm_id in seen_this_file:
            continue
        seen_this_file.add(crm_id)

        candidates = ClientRecord.query.filter_by(crm_id=crm_id).order_by(ClientRecord.id.desc()).all()
        was_paid = any(c.is_cleared or c.clawback_applied for c in candidates)
        client_rec = next((c for c in candidates if c.dropped_date), None)
        dropped_period = _period_of(_parse_date(client_rec.dropped_date)) if client_rec else None
        if not dropped_period or not was_paid:
            skipped_no_match.append(_client_label(row.get("client_name"), crm_id))
            continue

        amount = row.get("marketing_payout_debt") or 0.0
        fields = dict(
            agent_name=client_rec.agent_name,
            period_label=dropped_period,
            assigned_company=row.get("assigned_company") or "",
            enrolled_date=row.get("enrolled_date"),
            client_name=row.get("client_name") or client_rec.client_name,
            status=row.get("status") or "",
            marketing_payout_debt=amount,
            first_payment_cleared_date=row.get("first_payment_cleared_date"),
            pay_freq=row.get("pay_freq") or "",
            payments_made=row.get("payments_made"),
            marketing_payment_cleared=row.get("marketing_payment_cleared"),
            marketing_payment_chargeback=row.get("marketing_payment_chargeback"),
            file_dropped_date=row.get("file_dropped_date"),
            uploaded_filename=file.filename,
        )

        existing = existing_by_id.get(crm_id)
        if existing:
            for attr, value in fields.items():
                setattr(existing, attr, value)
            updated += 1
        else:
            db.session.add(CordobaChargebackEntry(crm_id=crm_id, **fields))
            listed += 1
        total += amount

    return listed, updated, round(total, 2), skipped_no_match


def _process_cordoba_file(file):
    """Parse one Cordoba payout file and apply both the paid-flag check (First Pays/EPF)
    and the chargeback-triggered agent clawback (Chargebacks tab). Unit-only crediting
    for low-value clients is no longer driven by this file's EPF tab (owner decision,
    July 2026) — that's now decided directly from the CRM export's own Credit Score
    column, see crm_parser.py. The EPF tab still feeds the "Cordoba Payout" confirmed-
    paid flag via _apply_cordoba_paid_flags."""
    file_bytes = file.read()
    parsed = parse_cordoba_payout(file_bytes)

    for err in parsed["errors"]:
        flash(f"{file.filename}: {err}", "error")

    new_count, flipped = _apply_cordoba_paid_flags(file, parsed)
    matched_count, unmatched_chargeback_ids = _mark_cordoba_chargeback_matches(file, parsed)
    (clawback_count, clawback_total, skipped_not_commissioned,
     skipped_not_confirmed_paid, skipped_already_clawed,
     skipped_no_dropped_date) = _apply_cordoba_chargebacks(file, parsed)
    listed_count, updated_count, listed_total, skipped_no_debt_match = _list_cordoba_chargebacks(file, parsed)

    db.session.commit()
    return (new_count, flipped, clawback_count, clawback_total,
            skipped_not_commissioned, skipped_not_confirmed_paid, skipped_already_clawed,
            skipped_no_dropped_date, matched_count, unmatched_chargeback_ids,
            listed_count, updated_count, listed_total, skipped_no_debt_match)


def _list_cordoba_uploads():
    """Groups every row across the four Cordoba ledger tables by
    uploaded_filename — the only "batch" identifier they share — into one
    summary per file, for the index page's "recent uploads" list."""
    filenames = set()
    for model in (CordobaPaidClient, CordobaChargedBackClient, CordobaChargebackMatchedClient, CordobaChargebackEntry):
        for row in db.session.query(model.uploaded_filename).distinct():
            if row[0]:
                filenames.add(row[0])

    summaries = []
    for filename in filenames:
        paid_rows = CordobaPaidClient.query.filter_by(uploaded_filename=filename).all()
        charged_back_rows = CordobaChargedBackClient.query.filter_by(uploaded_filename=filename).all()
        matched_rows = CordobaChargebackMatchedClient.query.filter_by(uploaded_filename=filename).all()
        entry_rows = CordobaChargebackEntry.query.filter_by(uploaded_filename=filename).all()
        timestamps = [r.uploaded_at for r in paid_rows + charged_back_rows + matched_rows + entry_rows if r.uploaded_at]
        summaries.append({
            "filename": filename,
            "uploaded_at": max(timestamps) if timestamps else None,
            "paid_count": len(paid_rows),
            "matched_count": len(matched_rows),
            "clawback_count": len(charged_back_rows),
            "clawback_total": round(sum(r.clawback_amount or 0.0 for r in charged_back_rows), 2),
            "listed_count": len(entry_rows),
        })

    summaries.sort(key=lambda s: s["uploaded_at"] or s["filename"], reverse=True)
    return summaries


def _delete_cordoba_upload(filename):
    """Reverses everything a Cordoba payout upload did, keyed by filename (the
    ledger tables' only shared "batch" identifier). Used by the index page's
    delete/reset action so a wrong file can be fully undone:

    - Actual clawback deductions (CordobaChargedBackClient rows) are reversed:
      the money is added back to the agent's commission for that period, the
      holding ClientRecord row _apply_cordoba_chargebacks created is removed,
      and if that leaves the agent with no units/clawback/clients left in the
      period (i.e. it was a pure zero-unit holding row created just to carry
      this clawback), the AgentCommission row — and the period itself, if it
      was left with no agents — is removed too.
    - cordoba_paid is unflagged on every ClientRecord for a crm_id, but ONLY if
      no OTHER Cordoba upload also confirmed that same crm_id (checked after
      this file's CordobaPaidClient rows are removed).
    - The two display-only ledgers (CordobaChargebackMatchedClient's "Cordoba
      Clawback: Yes" badge, CordobaChargebackEntry's reconciliation listing)
      are simply cleared for this filename.

    Returns a summary dict of what was reversed, for a flash message.
    """
    charged_back_rows = CordobaChargedBackClient.query.filter_by(uploaded_filename=filename).all()
    reversed_amount = 0.0

    for row in charged_back_rows:
        period = CommissionPeriod.query.filter_by(period_label=row.dropped_period).first()
        agent_row = (
            AgentCommission.query.filter_by(period_id=period.id, agent_name=row.agent_name).first()
            if period else None
        )

        if agent_row:
            client_rec = ClientRecord.query.filter_by(
                agent_commission_id=agent_row.id, crm_id=row.crm_id, clawback_applied=True,
            ).first()
            if client_rec:
                db.session.delete(client_rec)

            agent_row.clawback_amount = max(
                0.0, round((agent_row.clawback_amount or 0.0) - (row.clawback_amount or 0.0), 2)
            )
            agent_row.net_commission = max(0.0, round(agent_row.gross_commission - agent_row.clawback_amount, 2))
            note = f"Cordoba chargeback: -${row.clawback_amount:,.2f} for {row.client_name or row.crm_id} (ID {row.crm_id})"
            if agent_row.notes:
                agent_row.notes = " | ".join(p for p in agent_row.notes.split(" | ") if p != note)
            db.session.flush()

            remaining_clients = ClientRecord.query.filter_by(agent_commission_id=agent_row.id).count()
            if agent_row.units_cleared == 0 and agent_row.clawback_amount == 0 and remaining_clients == 0:
                db.session.delete(agent_row)
                db.session.flush()
                if period and AgentCommission.query.filter_by(period_id=period.id).count() == 0:
                    db.session.delete(period)

        reversed_amount += row.clawback_amount or 0.0
        db.session.delete(row)

    matched_removed = CordobaChargebackMatchedClient.query.filter_by(uploaded_filename=filename).delete()
    entries_removed = CordobaChargebackEntry.query.filter_by(uploaded_filename=filename).delete()

    # Used to be one SELECT + one UPDATE per crm_id in this file (a First Pays/EPF
    # roster is easily thousands of rows — thousands of individual round trips was
    # the single slowest part of deleting a Cordoba upload). Same fix shape as
    # ingest.py's bulk_delete_period: batch the "is this crm_id still confirmed by
    # some OTHER upload" check into one query, then unflag the rest in one UPDATE.
    paid_crm_ids = {
        r[0] for r in db.session.query(CordobaPaidClient.crm_id)
        .filter_by(uploaded_filename=filename)
    }
    CordobaPaidClient.query.filter_by(uploaded_filename=filename).delete(synchronize_session=False)
    db.session.flush()

    still_confirmed = {
        r[0] for r in db.session.query(CordobaPaidClient.crm_id)
        .filter(CordobaPaidClient.crm_id.in_(paid_crm_ids))
    } if paid_crm_ids else set()
    to_unflag = paid_crm_ids - still_confirmed
    unflagged = ClientRecord.query.filter(ClientRecord.crm_id.in_(to_unflag)).update(
        {"cordoba_paid": False}, synchronize_session=False,
    ) if to_unflag else 0

    db.session.commit()
    return {
        "clawbacks_reversed": len(charged_back_rows),
        "amount_reversed": round(reversed_amount, 2),
        "matched_removed": matched_removed,
        "entries_removed": entries_removed,
        "paid_confirmations_removed": len(paid_crm_ids),
        "cordoba_paid_unflagged": unflagged,
    }


@bp.route("/upload-cordoba-payout", methods=["POST"])
def upload_cordoba_payout():
    """Upload one or more Cordoba payout files (.xlsx): First Pays/EPF flag confirmed
    payouts, Chargebacks tab triggers agent clawbacks for previously-paid clients."""
    files = [f for f in request.files.getlist("cordoba_file") if f and f.filename]
    if not files:
        flash("No file selected.", "error")
        return redirect(url_for("main.index"))

    bad_names = [f.filename for f in files if not _allowed_xlsx_file(f.filename)]
    if bad_names:
        flash(f"Only .xlsx files are accepted for Cordoba payout uploads: {', '.join(bad_names)}", "error")
        return redirect(url_for("main.index"))

    results = [_process_cordoba_file(file) for file in files]
    new_total = sum(r[0] for r in results)
    flipped_total = sum(r[1] for r in results)
    clawback_count_total = sum(r[2] for r in results)
    clawback_amount_total = sum(r[3] for r in results)
    skipped_not_commissioned = [name for r in results for name in r[4]]
    skipped_not_confirmed_paid = [name for r in results for name in r[5]]
    skipped_already_clawed = [name for r in results for name in r[6]]
    skipped_no_dropped_date = [name for r in results for name in r[7]]
    matched_total = sum(r[8] for r in results)
    unmatched_chargeback_ids = [name for r in results for name in r[9]]
    listed_total = sum(r[10] for r in results)
    updated_total = sum(r[11] for r in results)
    listed_amount_total = sum(r[12] for r in results)
    skipped_no_debt_match = [name for r in results for name in r[13]]

    file_word = "file" if len(files) == 1 else f"{len(files)} files"
    flash(
        f"Cordoba payout processed ({file_word}): {new_total} newly recorded ID(s) in the ledger, "
        f"{flipped_total} client record(s) marked Cordoba Payout = Yes.",
        "success",
    )
    if matched_total > 0:
        flash(
            f"Cordoba chargebacks: {matched_total} client(s) matched in our commission reports "
            f"and marked \"Cordoba Clawback: Yes\".",
            "success",
        )
    if clawback_count_total > 0:
        flash(
            f"Cordoba chargebacks: {clawback_count_total} client(s) charged back, "
            f"${clawback_amount_total:,.2f} clawed back from agent commissions.",
            "success",
        )
    if listed_total > 0 or updated_total > 0:
        parts = []
        if listed_total > 0:
            parts.append(f"{listed_total} newly listed")
        if updated_total > 0:
            parts.append(f"{updated_total} refreshed with this file's numbers")
        flash(
            f"Cordoba Charge back: {', '.join(parts)} (${listed_amount_total:,.2f} total "
            f"Marketing Payout Debt this upload) on the relevant agents' commission reports "
            f"for reference — informational only, not deducted.",
            "success",
        )

    def _flash_skipped(names, reason):
        if not names:
            return
        shown = ", ".join(names[:10])
        more = f" and {len(names) - 10} more" if len(names) > 10 else ""
        flash(f"{len(names)} charged-back client(s) {reason}: {shown}{more}.", "error")

    _flash_skipped(skipped_not_commissioned, "were never recorded as commissioned here — no clawback applied")
    _flash_skipped(skipped_not_confirmed_paid,
                   "were never confirmed paid via a First Pays/EPF upload — no clawback applied")
    _flash_skipped(skipped_already_clawed,
                   "were already clawed back via a CRM upload or history import — not deducted twice")
    _flash_skipped(skipped_no_dropped_date,
                   "have no Dropped Date recorded in our own CRM data yet — upload a CRM export "
                   "reflecting the drop, then re-upload this Chargebacks file")
    _flash_skipped(unmatched_chargeback_ids, "were not found in any of our commission reports — no match")
    _flash_skipped(skipped_no_debt_match,
                   "were not found in our commission reports, have no Dropped Date on file yet, or were "
                   "never actually paid commission (dropped before payout) — not listed under "
                   "\"Cordoba Charge back\" on any agent's report")
    return redirect(url_for("main.index"))


def _save_commission_history_period(period_label, results, filename, already_cordoba_paid_ids):
    """Save one month's worth of parsed historical-ledger results as a real
    CommissionPeriod + AgentCommission + ClientRecord rows, same shape the CRM flow
    produces, so this history is indistinguishable from a real upload for the purposes
    of Cordoba chargeback matching (_apply_cordoba_chargebacks looks up
    ClientRecord.is_cleared=True by crm_id, regardless of which upload flow created it)."""
    period = CommissionPeriod(period_label=period_label, filename=filename, total_agents=len(results))
    db.session.add(period)
    db.session.flush()

    for r in results:
        cleared_clients = r.pop("_cleared_clients", [])
        clawback_clients = r.pop("_clawback_clients", [])

        agent_obj = AgentCommission(period_id=period.id, **r)
        db.session.add(agent_obj)
        db.session.flush()

        for cr in cleared_clients:
            db.session.add(_new_client_record(
                period.id, agent_obj.id, cr,
                is_cleared=True,
                is_pending=False,
                is_cancelled=False,
                commission_on_client=round(cr.get("enrolled_debt", 0.0) * agent_obj.tier_rate, 2),
                cordoba_paid=cr.get("crm_id") in already_cordoba_paid_ids,
            ))

        for cr in clawback_clients:
            db.session.add(_new_client_record(
                period.id, agent_obj.id, cr,
                is_cleared=False,
                is_pending=False,
                is_cancelled=True,
                commission_on_client=0.0,
                clawback_applied=True,
                clawback_period_id=period.id,
                clawback_amount=cr.get("clawback_amount", 0.0),
            ))

    return period


@bp.route("/upload-commission-history", methods=["POST"])
def upload_commission_history():
    """Backfill past commission history from a prior account manager's ledger (.xlsx or
    .csv, NOT a CRM export — see commission_history_parser.py for the expected columns).
    Recreates real CommissionPeriod/AgentCommission/ClientRecord rows for those months
    so a later Cordoba Chargebacks-tab upload can find and claw back agents who were
    paid on a client before this app existed."""
    files = [f for f in request.files.getlist("history_file") if f and f.filename]
    if not files:
        flash("No file selected.", "error")
        return redirect(url_for("main.index"))

    bad_names = [f.filename for f in files if not _allowed_history_file(f.filename)]
    if bad_names:
        flash(f"Only .xlsx or .csv files are accepted for commission history uploads: {', '.join(bad_names)}", "error")
        return redirect(url_for("main.index"))

    year_raw = (request.form.get("history_year") or "").strip()
    if not year_raw.isdigit():
        flash("Please enter a valid year for the commission history file (the Month column has no year).", "error")
        return redirect(url_for("main.index"))
    year = int(year_raw)

    already_cordoba_paid_ids = {p.crm_id for p in CordobaPaidClient.query.all()}

    saved_period_ids = []
    total_periods_skipped = 0

    for file in files:
        file_bytes = file.read()
        parsed = parse_commission_history(file_bytes, file.filename, year)

        for err in parsed["errors"]:
            flash(f"{file.filename}: {err}", "error")

        for period_data in parsed["periods"]:
            period_label = period_data["period_label"]
            existing = CommissionPeriod.query.filter_by(period_label=period_label).first()
            if existing:
                flash(
                    f"Period {period_label} already exists (uploaded {existing.uploaded_at.strftime('%Y-%m-%d')}). "
                    "Delete it first before re-importing history for that month.", "error",
                )
                total_periods_skipped += 1
                continue

            period = _save_commission_history_period(
                period_label, period_data["results"], file.filename, already_cordoba_paid_ids
            )
            saved_period_ids.append(period.id)

        db.session.commit()

    if saved_period_ids:
        flash(f"Commission history import: {len(saved_period_ids)} month(s) backfilled.", "success")
    if total_periods_skipped:
        flash(f"{total_periods_skipped} month(s) skipped because a period already existed.", "error")

    if len(saved_period_ids) == 1:
        return redirect(url_for("main.period_detail", period_id=saved_period_ids[0]))
    return redirect(url_for("main.history"))


@bp.route("/period/<int:period_id>")
def period_detail(period_id):
    period = CommissionPeriod.query.get_or_404(period_id)
    agents = AgentCommission.query.filter_by(period_id=period_id).order_by(AgentCommission.agent_name).all()

    total_net = sum(a.net_commission for a in agents)
    total_gross = sum(a.gross_commission for a in agents)
    total_clawback = sum(a.clawback_amount for a in agents)
    bonus_eligible = sum(1 for a in agents if a.quality_bonus_eligible)
    penalty_count = sum(1 for a in agents if a.cancellation_penalty_applied)
    nsf_count = sum(1 for a in agents if a.nsf_flagged)
    pending_count = sum(1 for a in agents if a.pending_units > 0)

    return render_template(
        "results.html",
        period=period,
        agents=agents,
        total_net=total_net,
        total_gross=total_gross,
        total_clawback=total_clawback,
        bonus_eligible=bonus_eligible,
        penalty_count=penalty_count,
        nsf_count=nsf_count,
        pending_count=pending_count,
        units_to_next_tier_map=_units_to_next_tier_map(agents),
    )


@bp.route("/period/<int:period_id>/agent/<int:agent_id>")
def agent_detail(period_id, agent_id):
    period = CommissionPeriod.query.get_or_404(period_id)
    agent = AgentCommission.query.get_or_404(agent_id)
    clients = ClientRecord.query.filter_by(agent_commission_id=agent_id).all()
    clawback_clients = [c for c in clients if c.clawback_applied]
    active_clients = [c for c in clients if not c.clawback_applied]

    # Per-client "Cordoba Clawback" flag for the Cleared Clients table — looked up from
    # CordobaChargebackMatchedClient (crm_id matched a Chargebacks-tab row against ANY
    # of our commission reports), not from the money ledger (CordobaChargedBackClient).
    # Owner policy (confirmed July 2026): shows Yes as soon as the client is recognized,
    # even if the actual dollar deduction is still blocked on a gate in
    # _apply_cordoba_chargebacks (most commonly: no Dropped Date on file yet). Purely
    # informational — does not affect tier, units, or commission.
    crm_ids = {c.crm_id for c in active_clients if c.crm_id}
    cordoba_charged_back_ids = {
        cb.crm_id for cb in
        CordobaChargebackMatchedClient.query.filter(CordobaChargebackMatchedClient.crm_id.in_(crm_ids)).all()
    } if crm_ids else set()

    # Informational-only "Cordoba Charge back" line items for this agent in this period
    # (matched by dropped month) — see _list_cordoba_chargebacks. Never affects
    # gross_commission/net_commission/clawback_amount above.
    cordoba_chargeback_entries = CordobaChargebackEntry.query.filter_by(
        agent_name=agent.agent_name, period_label=period.period_label,
    ).order_by(CordobaChargebackEntry.uploaded_at).all()

    return render_template(
        "agent_detail.html",
        period=period,
        agent=agent,
        clients=active_clients,
        clawback_clients=clawback_clients,
        cordoba_charged_back_ids=cordoba_charged_back_ids,
        cordoba_chargeback_entries=cordoba_chargeback_entries,
        units_to_next_tier=units_to_next_tier(agent.units_cleared, agent.agent_name),
        commission_gain_at_next_tier=commission_gain_at_next_tier(
            agent.adjusted_tier, agent.total_cleared_debt, agent.gross_commission, agent.agent_name,
        ),
    )


@bp.route("/period/<int:period_id>/export")
def export_period(period_id):
    period = CommissionPeriod.query.get_or_404(period_id)
    agents = AgentCommission.query.filter_by(period_id=period_id).order_by(AgentCommission.agent_name).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Agent Name", "Units Cleared", "Cleared Debt", "Cancel Rate %",
        "Raw Tier", "Adjusted Tier", "Rate %",
        "Gross Commission", "Clawback", "Net Commission",
        "Quality Bonus Eligible", "Cancel Penalty Applied",
        "NSF Flagged", "Pending Units", "Pending Debt", "Notes",
    ])
    for a in agents:
        writer.writerow([
            a.agent_name, a.units_cleared, f"{a.total_cleared_debt:.2f}",
            f"{a.cancellation_rate:.1f}",
            a.raw_tier, a.adjusted_tier, f"{a.tier_rate*100:.2f}",
            f"{a.gross_commission:.2f}", f"{a.clawback_amount:.2f}", f"{a.net_commission:.2f}",
            "Yes" if a.quality_bonus_eligible else "No",
            "Yes" if a.cancellation_penalty_applied else "No",
            "Yes" if a.nsf_flagged else "No",
            a.pending_units, f"{a.pending_debt:.2f}",
            a.notes,
        ])

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=commissions_{period.period_label}.csv"},
    )


CLIENT_EXPORT_COLUMNS = [
    "Type", "ID", "Client Name", "Enrolled Date", "Enrolled Debt", "Status",
    "1st Payment Cleared Date", "2nd Payment Cleared Date", "Dropped Date",
    "Payments Made", "Pay Freq.", "# NSF", "Credit Score",
    "Commission on Client", "Clawback Amount", "Cordoba Payout", "Cordoba Clawback",
]


def _client_export_rows(clients, cordoba_charged_back_ids=frozenset()):
    clawback_clients = [c for c in clients if c.clawback_applied]
    active_clients = [c for c in clients if not c.clawback_applied]
    rows = []
    for c in active_clients:
        t = "Cleared" if c.is_cleared else ("Pending" if c.is_pending else "Cancelled")
        rows.append([
            t, c.crm_id or "", c.client_name, c.enrolled_date or "",
            f"{c.enrolled_debt:.2f}", c.status,
            c.first_payment_cleared_date, c.second_payment_cleared_date or "",
            c.dropped_date or "",
            c.payments_made, c.pay_freq or "", c.nsf_count,
            c.credit_score if c.credit_score is not None else "",
            f"{c.commission_on_client:.2f}", "",
            ("Yes" if c.cordoba_paid else "No") if c.is_cleared else "",
            ("Yes" if c.crm_id in cordoba_charged_back_ids else "No") if c.is_cleared else "",
        ])
    for c in clawback_clients:
        rows.append([
            "Clawback", c.crm_id or "", c.client_name, c.enrolled_date or "",
            f"{c.enrolled_debt:.2f}", c.status,
            c.first_payment_cleared_date, c.second_payment_cleared_date or "",
            c.dropped_date or "",
            c.payments_made, c.pay_freq or "", c.nsf_count,
            c.credit_score if c.credit_score is not None else "",
            "", f"-{c.clawback_amount:.2f}", "", "",
        ])
    return rows


CORDOBA_CHARGEBACK_EXPORT_COLUMNS = [
    "Assigned Company", "Enrolled Date", "ID", "Full Name", "Status",
    "Marketing Payout Debt", "1st Payment Cleared Date", "Pay Freq.", "Payments Made",
    "Marketing Payment Cleared", "Marketing Payment Chargeback", "Dropped Date",
]


def _write_cordoba_chargeback_block(writer, entries, agent_name=None):
    """Writes a "Cordoba Charge back" block into a CSV export, verbatim in the same
    column shape as the Chargebacks tab itself (see CordobaChargebackEntry / owner
    request July 2026) — a separate mini-table (title, header, rows) rather than
    columns woven into CLIENT_EXPORT_COLUMNS, so it reads exactly like the source file
    for reconciliation. Purely informational — none of this feeds Clawback Amount.
    agent_name is included in the title when writing into the combined all-agents
    export, where a bare "Cordoba Charge back" title wouldn't say whose block it is."""
    if not entries:
        return
    writer.writerow([])
    writer.writerow([f"Cordoba Charge back — {agent_name}" if agent_name else "Cordoba Charge back"])
    writer.writerow(CORDOBA_CHARGEBACK_EXPORT_COLUMNS)
    for e in entries:
        writer.writerow([
            e.assigned_company or "", e.enrolled_date or "", e.crm_id or "",
            e.client_name or "", e.status or "",
            f"${e.marketing_payout_debt:,.2f}",
            e.first_payment_cleared_date or "", e.pay_freq or "",
            e.payments_made if e.payments_made is not None else "",
            e.marketing_payment_cleared or "", e.marketing_payment_chargeback or "",
            e.file_dropped_date or "",
        ])


@bp.route("/period/<int:period_id>/agent/<int:agent_id>/export")
def export_agent(period_id, agent_id):
    period = CommissionPeriod.query.get_or_404(period_id)
    agent = AgentCommission.query.get_or_404(agent_id)
    clients = ClientRecord.query.filter_by(agent_commission_id=agent_id).all()
    crm_ids = {c.crm_id for c in clients if c.crm_id}
    cordoba_charged_back_ids = {
        cb.crm_id for cb in
        CordobaChargebackMatchedClient.query.filter(CordobaChargebackMatchedClient.crm_id.in_(crm_ids)).all()
    } if crm_ids else set()

    cordoba_chargeback_entries = CordobaChargebackEntry.query.filter_by(
        agent_name=agent.agent_name, period_label=period.period_label,
    ).order_by(CordobaChargebackEntry.uploaded_at).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(CLIENT_EXPORT_COLUMNS)
    for row in _client_export_rows(clients, cordoba_charged_back_ids):
        writer.writerow(row)
    _write_cordoba_chargeback_block(writer, cordoba_chargeback_entries)

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={agent.agent_name.replace(' ','_')}_{period.period_label}.csv"},
    )


def _period_display_label(period_label):
    """Formats the stored "YYYY-MM" period_label as "June 2026" for display in
    exports. Falls back to the raw label if it's ever in an unexpected shape."""
    try:
        return datetime.strptime(period_label, "%Y-%m").strftime("%B %Y")
    except (ValueError, TypeError):
        return period_label


def _safe_sheet_title(name, used_titles):
    """Excel sheet titles: max 31 chars, and : \\ / ? * [ ] are illegal — dedup
    on collision (e.g. two agents sharing a truncated name)."""
    cleaned = re.sub(r'[:\\/?*\[\]]', "", name or "").strip() or "Agent"
    base = cleaned[:31]
    title = base
    n = 2
    while title in used_titles:
        suffix = f" ({n})"
        title = base[:31 - len(suffix)] + suffix
        n += 1
    used_titles.add(title)
    return title


CURRENCY_NUMBER_FORMAT = "$#,##0.00"
# Columns within a full ["Agent Name", "Tier", "Rate %"] + CLIENT_EXPORT_COLUMNS row
# that should be written as real numeric currency cells, not text.
_CURRENCY_COLUMN_INDICES = [
    3 + CLIENT_EXPORT_COLUMNS.index("Enrolled Debt"),
    3 + CLIENT_EXPORT_COLUMNS.index("Commission on Client"),
]

CORDOBA_HEADER_FILL = PatternFill(start_color="FFC00000", end_color="FFC00000", fill_type="solid")
CORDOBA_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
CORDOBA_CHARGEBACK_XLSX_COLUMNS = ["Agent Name"] + CORDOBA_CHARGEBACK_EXPORT_COLUMNS


def _write_agent_client_rows(ws, agent, clients, cordoba_charged_back_ids, period_display_label):
    """Appends this agent's client-detail rows into ws. Enrolled Debt and Commission
    on Client are written as real numeric cells with a currency number format,
    instead of the plain formatted-string values _client_export_rows returns (which
    are fine for CSV but would show as text, not dollar amounts, in Excel). The
    Commission Period column is appended at the end of each row rather than inserted
    among CLIENT_EXPORT_COLUMNS, so _CURRENCY_COLUMN_INDICES (computed off that list's
    original positions) stays valid."""
    for row in _client_export_rows(clients, cordoba_charged_back_ids):
        full_row = [agent.agent_name, agent.adjusted_tier, f"{agent.tier_rate*100:.2f}"] + row + [period_display_label]
        for idx in _CURRENCY_COLUMN_INDICES:
            if full_row[idx] not in ("", None):
                full_row[idx] = float(full_row[idx])
        ws.append(full_row)
        r = ws.max_row
        for idx in _CURRENCY_COLUMN_INDICES:
            if isinstance(full_row[idx], float):
                ws.cell(row=r, column=idx + 1).number_format = CURRENCY_NUMBER_FORMAT


def _write_cordoba_header(ws):
    """Writes the Cordoba Charge back header row — Agent Name plus the Chargebacks-tab
    columns, plus a trailing Commission Period column — as solid red with white bold
    text."""
    header = CORDOBA_CHARGEBACK_XLSX_COLUMNS + ["Commission Period"]
    ws.append(header)
    # ws.max_row only advances once a row holds a real value, so it's always safe to
    # read right after appending an actually-populated row like this header.
    header_row_num = ws.max_row
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=header_row_num, column=col)
        cell.fill = CORDOBA_HEADER_FILL
        cell.font = CORDOBA_HEADER_FONT


def _write_cordoba_rows(ws, entries, agent_name, period_display_label):
    """Appends this agent's Cordoba Charge back data rows (no header) into ws.
    Marketing Payout Debt is a real numeric currency cell (not a formatted string) so
    the Dashboard Summary tab's Sum of To Subtract formula can SUMIF over it. The
    Commission Period column is appended at the end so column G (Marketing Payout
    Debt), referenced by that SUMIF, stays put."""
    for e in entries:
        ws.append([
            agent_name,
            e.assigned_company or "", e.enrolled_date or "", e.crm_id or "",
            e.client_name or "", e.status or "",
            e.marketing_payout_debt or 0.0,
            e.first_payment_cleared_date or "", e.pay_freq or "",
            e.payments_made if e.payments_made is not None else "",
            e.marketing_payment_cleared or "", e.marketing_payment_chargeback or "",
            e.file_dropped_date or "",
            period_display_label,
        ])
        ws.cell(row=ws.max_row, column=7).number_format = CURRENCY_NUMBER_FORMAT


DASHBOARD_TITLE_FONT = Font(bold=True, size=16)
DASHBOARD_HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
DASHBOARD_HEADER_FONT = Font(bold=True, italic=True)
DASHBOARD_NEGATIVE_CURRENCY_FORMAT = '$#,##0.00;($#,##0.00);"$ -"'


def _write_dashboard_summary(ws, agent_rows, chargeback_sheet_title, period_display_label):
    """Writes the two-table period overview: "Enrolled debt by Rep" (debt/units/
    RevShares — units where Credit Score <= 500, per owner — per agent) and
    "Commission payout" (rate/bonus/net commission per agent). Bonus is left blank
    for manual entry — the app has no dollar bonus calculation (quality_bonus_eligible
    is display-only). Sum of To Subtract, Total Units, and every Grand Total cell are
    live Excel formulas (not precomputed numbers) so editing a value in this sheet, or
    a Marketing Payout Debt value on the chargeback_sheet_title tab, recalculates the
    totals automatically — per owner request."""
    LEFT_COL = 1     # A: Sales Rep
    # Left table spans A-G (7 columns, incl. trailing Commission Period); H is a
    # blank gap; right table starts at I.
    RIGHT_COL = 9

    ws.cell(row=1, column=LEFT_COL, value="Enrolled debt by Rep").font = DASHBOARD_TITLE_FONT
    ws.cell(row=1, column=RIGHT_COL, value="Commission payout").font = DASHBOARD_TITLE_FONT

    left_headers = ["Sales Rep", "Sum of Enrolled Debt", "Sum of To Subtract", "Sum of Units",
                     "RevShares", "Total Units", "Commission Period"]
    right_headers = ["Sales Rep", "Rate %", "Bonus", "Total Commissions"]

    header_row = 2
    for i, h in enumerate(left_headers):
        c = ws.cell(row=header_row, column=LEFT_COL + i, value=h)
        c.font = DASHBOARD_HEADER_FONT
        c.fill = DASHBOARD_HEADER_FILL
    for i, h in enumerate(right_headers):
        c = ws.cell(row=header_row, column=RIGHT_COL + i, value=h)
        c.font = DASHBOARD_HEADER_FONT
        c.fill = DASHBOARD_HEADER_FILL

    agent_col = get_column_letter(LEFT_COL)
    units_col = get_column_letter(LEFT_COL + 3)
    revshares_col = get_column_letter(LEFT_COL + 4)
    rate_col = get_column_letter(RIGHT_COL + 1)
    cb_agent_col = "A"       # Agent Name column on the chargeback sheet
    cb_debt_col = "G"        # Marketing Payout Debt column on the chargeback sheet

    first_data_row = header_row + 1
    r = first_data_row
    for row_data in agent_rows:
        ws.cell(row=r, column=LEFT_COL, value=row_data["agent_name"])
        ws.cell(row=r, column=LEFT_COL + 1, value=row_data["enrolled_debt"]).number_format = CURRENCY_NUMBER_FORMAT
        # Sum of To Subtract = SUM(Marketing Payout Debt) for this agent's chargeback
        # rows x their rate% — sourced live from the chargeback tab and this row's own
        # Rate % cell, so either one changing recalculates this automatically.
        to_subtract_cell = ws.cell(
            row=r, column=LEFT_COL + 2,
            value=(f"=-SUMIF('{chargeback_sheet_title}'!${cb_agent_col}:${cb_agent_col},"
                   f"{agent_col}{r},'{chargeback_sheet_title}'!${cb_debt_col}:${cb_debt_col})"
                   f"*{rate_col}{r}"),
        )
        to_subtract_cell.number_format = DASHBOARD_NEGATIVE_CURRENCY_FORMAT
        ws.cell(row=r, column=LEFT_COL + 3, value=row_data["units"])
        ws.cell(row=r, column=LEFT_COL + 4, value=row_data["revshares"] or None)
        # Total Units = Sum of Units + RevShares, live — editing either recalculates it.
        ws.cell(row=r, column=LEFT_COL + 5, value=f"={units_col}{r}+{revshares_col}{r}")
        ws.cell(row=r, column=LEFT_COL + 6, value=period_display_label)

        ws.cell(row=r, column=RIGHT_COL, value=row_data["agent_name"])
        ws.cell(row=r, column=RIGHT_COL + 1, value=row_data["rate"]).number_format = "0.00%"
        # Bonus (RIGHT_COL + 2) intentionally left blank for manual entry.
        ws.cell(row=r, column=RIGHT_COL + 3, value=row_data["net_commission"]).number_format = CURRENCY_NUMBER_FORMAT
        r += 1

    last_data_row = r - 1
    total_row = r

    ws.cell(row=total_row, column=LEFT_COL, value="Grand Total").font = Font(bold=True)
    if last_data_row >= first_data_row:
        for offset, fmt in ((1, CURRENCY_NUMBER_FORMAT), (2, DASHBOARD_NEGATIVE_CURRENCY_FORMAT),
                             (3, None), (4, None), (5, None)):
            col_letter = get_column_letter(LEFT_COL + offset)
            cell = ws.cell(row=total_row, column=LEFT_COL + offset,
                            value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})")
            cell.font = Font(bold=True)
            if fmt:
                cell.number_format = fmt

        commission_col = get_column_letter(RIGHT_COL + 3)
        commission_cell = ws.cell(
            row=total_row, column=RIGHT_COL + 3,
            value=f"=SUM({commission_col}{first_data_row}:{commission_col}{last_data_row})",
        )
        commission_cell.font = Font(bold=True)
        commission_cell.number_format = CURRENCY_NUMBER_FORMAT

    ws.cell(row=total_row, column=RIGHT_COL + 2, value="Total Commissions:").font = Font(bold=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 18


@bp.route("/period/<int:period_id>/export-by-agent")
def export_by_agent(period_id):
    """One .xlsx workbook for the whole period: a "Dashboard Summary" sheet (enrolled
    debt/units/clawback and commission payout per agent), an "All Agents" sheet
    combining every agent's cleared/pending/cancelled client rows (no chargebacks
    mixed in, per owner request), an "All Chargeback" sheet combining every agent's
    Cordoba Charge back rows, then one sheet per agent with their own client rows AND
    their own Cordoba Charge back block appended below — same underlying data as
    export_all_agents but split into tabs (matches the agent_client_details_by_agent
    style)."""
    period = CommissionPeriod.query.get_or_404(period_id)
    agents = AgentCommission.query.filter_by(period_id=period_id).order_by(AgentCommission.agent_name).all()
    period_display_label = _period_display_label(period.period_label)

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles = set()

    dashboard_ws = workbook.create_sheet(_safe_sheet_title("Dashboard Summary", used_titles))

    combined_ws = workbook.create_sheet(_safe_sheet_title("All Agents", used_titles))
    combined_ws.append(["Agent Name", "Tier", "Rate %"] + CLIENT_EXPORT_COLUMNS + ["Commission Period"])

    chargeback_ws = workbook.create_sheet(_safe_sheet_title("All Chargeback", used_titles))
    _write_cordoba_header(chargeback_ws)

    dashboard_rows = []

    for agent in agents:
        ws = workbook.create_sheet(_safe_sheet_title(agent.agent_name, used_titles))
        ws.append(["Agent Name", "Tier", "Rate %"] + CLIENT_EXPORT_COLUMNS + ["Commission Period"])

        clients = ClientRecord.query.filter_by(agent_commission_id=agent.id).all()
        crm_ids = {c.crm_id for c in clients if c.crm_id}
        cordoba_charged_back_ids = {
            cb.crm_id for cb in
            CordobaChargebackMatchedClient.query.filter(CordobaChargebackMatchedClient.crm_id.in_(crm_ids)).all()
        } if crm_ids else set()

        _write_agent_client_rows(ws, agent, clients, cordoba_charged_back_ids, period_display_label)
        _write_agent_client_rows(combined_ws, agent, clients, cordoba_charged_back_ids, period_display_label)

        revshares = sum(1 for c in clients if c.is_cleared and c.is_low_credit)
        dashboard_rows.append({
            "agent_name": agent.agent_name,
            "enrolled_debt": agent.total_cleared_debt,
            "units": agent.units_cleared - revshares,
            "revshares": revshares,
            "rate": agent.tier_rate,
            "net_commission": agent.net_commission,
        })

        cordoba_chargeback_entries = CordobaChargebackEntry.query.filter_by(
            agent_name=agent.agent_name, period_label=period.period_label,
        ).order_by(CordobaChargebackEntry.uploaded_at).all()

        if cordoba_chargeback_entries:
            ws.append([])
            _write_cordoba_header(ws)
            _write_cordoba_rows(ws, cordoba_chargeback_entries, agent.agent_name, period_display_label)

        _write_cordoba_rows(chargeback_ws, cordoba_chargeback_entries, agent.agent_name, period_display_label)

    _write_dashboard_summary(dashboard_ws, dashboard_rows, chargeback_ws.title, period_display_label)

    if not agents:
        workbook.create_sheet("No Agents")

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
                f"attachment; filename=agent_client_details_by_agent_{period.period_label}.xlsx"
        },
    )


@bp.route("/period/<int:period_id>/delete", methods=["POST"])
def delete_period(period_id):
    period = CommissionPeriod.query.get_or_404(period_id)
    db.session.delete(period)
    db.session.commit()
    flash(f"Period {period.period_label} deleted.", "success")
    return redirect(url_for("main.history"))


@bp.route("/uploads/crm/delete", methods=["POST"])
def delete_crm_upload():
    """Deletes every period a single CRM-export upload created (it can span
    several months in one file) in one action, instead of one period at a
    time from the period detail page."""
    filename = request.form.get("filename") or ""
    deleted = _delete_periods_by_filename(filename, CRM_SOURCES)
    if deleted:
        flash(f"Deleted {len(deleted)} period(s) from \"{filename}\": {', '.join(deleted)}. "
              "Re-upload the CRM export to re-import them.", "success")
    else:
        flash(f"No CRM-export periods found for \"{filename}\".", "error")
    return redirect(url_for("main.index"))


@bp.route("/uploads/history/delete", methods=["POST"])
def delete_history_upload():
    """Deletes every period a single commission-history-backfill upload
    created, in one action."""
    filename = request.form.get("filename") or ""
    deleted = _delete_periods_by_filename(filename, HISTORY_SOURCES)
    if deleted:
        flash(f"Deleted {len(deleted)} backfilled period(s) from \"{filename}\": {', '.join(deleted)}. "
              "Re-upload the history file to re-import them.", "success")
    else:
        flash(f"No commission-history periods found for \"{filename}\".", "error")
    return redirect(url_for("main.index"))


@bp.route("/uploads/cordoba/delete", methods=["POST"])
def delete_cordoba_upload_route():
    """Fully reverses one Cordoba payout upload: reinstates any clawed-back
    commission, un-flags Cordoba Payout confirmations no other file also
    confirmed, and clears the two display-only ledgers. See
    _delete_cordoba_upload for exactly what this undoes."""
    filename = request.form.get("filename") or ""
    result = _delete_cordoba_upload(filename)
    if result["clawbacks_reversed"]:
        flash(
            f"Reversed {result['clawbacks_reversed']} clawback(s) from \"{filename}\" "
            f"(${result['amount_reversed']:,.2f} restored to agent commissions).", "success",
        )
    flash(
        f"\"{filename}\" reset: {result['paid_confirmations_removed']} paid confirmation(s) removed "
        f"({result['cordoba_paid_unflagged']} client record(s) un-flagged), "
        f"{result['matched_removed']} chargeback match(es) and {result['entries_removed']} "
        "reconciliation entry(ies) cleared.", "success",
    )
    return redirect(url_for("main.index"))


@bp.route("/history")
def history():
    periods = CommissionPeriod.query.order_by(CommissionPeriod.uploaded_at.desc()).all()
    return render_template("history.html", periods=periods)
