"""
Parses the Cordoba (funder) payout export (.xlsx). The file has 3 tabs:

- First Pays / EPF: confirm Cordoba actually paid us on a client — checked against
  our own ClientRecord IDs to flag cordoba_paid = True (purely informational).
- Chargebacks: confirms Cordoba took a marketing payout BACK from us on a client.
  Returned as raw rows here; routes.py cross-references each ID against our own
  ClientRecord history (this tab has no agent/rep column of its own) and claws back
  the agent's commission if we ever paid them on that client.
"""

import io
from datetime import date, datetime

import openpyxl

REQUIRED_SHEETS = {"first pays", "epf"}


def _clean_id(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def _clean_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%m/%d/%Y")
    return str(value).strip() or None


def _clean_amount(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace("$", "").replace(",", "") or 0)
    except ValueError:
        return 0.0


def _sheet_by_name(workbook, wanted_lower: str):
    for name in workbook.sheetnames:
        if name.strip().lower() == wanted_lower:
            return workbook[name]
    return None


def _header_map(sheet) -> dict:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return {str(h).strip().lower(): idx for idx, h in enumerate(header_row) if h}


def _parse_chargebacks(workbook, errors: list) -> list:
    """Owner policy (confirmed July 2026): only the 'ID' column matters for the actual
    clawback deduction — it's cross-referenced against our own ClientRecord history in
    routes.py. The client's debt and the dropped-date-to-place-the-deduction still come
    from OUR OWN records, not this file. 'Full Name' is read only to make skip/flash
    messages readable.

    'Marketing Payout Debt' (owner request, July 2026) is also read here, but ONLY for
    the separate, display-only "listed at the bottom of the agent's commission report"
    feature in routes.py (_list_cordoba_marketing_payout_debt / CordobaMarketingPayoutDebtEntry)
    — it is never used for the real clawback math above, which still recalculates via
    calculate_clawback_amount on our own enrolled_debt."""
    sheet = _sheet_by_name(workbook, "chargebacks")
    if sheet is None:
        return []

    cols = _header_map(sheet)
    if "id" not in cols:
        errors.append("Chargebacks tab is missing an 'ID' column.")
        return []

    def cell(row, key):
        idx = cols.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        crm_id = _clean_id(cell(row, "id"))
        if not crm_id:
            continue
        rows.append({
            "crm_id": crm_id,
            "client_name": cell(row, "full name") or "",
            "marketing_payout_debt": _clean_amount(cell(row, "marketing payout debt")),
        })
    return rows


def parse_cordoba_payout(file_bytes: bytes) -> dict:
    """
    Returns:
    {
        "paid_ids": [ {"crm_id": str, "client_name": str, "source": "first_pays"|"epf"}, ... ],
        "chargebacks": [ {"crm_id": str, "client_name": str}, ... ],
        "errors": [str, ...],
    }
    """
    errors = []

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        return {"paid_ids": [], "chargebacks": [],
                "errors": ["Could not read the file — expected an .xlsx "
                           "workbook with First Pays and EPF tabs."]}

    sheet_names_lower = {name.strip().lower() for name in workbook.sheetnames}
    missing = REQUIRED_SHEETS - sheet_names_lower
    if missing:
        errors.append(f"Missing tab(s) in Cordoba file: {', '.join(sorted(missing))}")

    paid_ids = []

    first_pays = _sheet_by_name(workbook, "first pays")
    if first_pays is not None:
        cols = _header_map(first_pays)
        if "id" not in cols:
            errors.append("First Pays tab is missing an 'ID' column.")
        else:
            for row in first_pays.iter_rows(min_row=2, values_only=True):
                crm_id = _clean_id(row[cols["id"]]) if cols["id"] < len(row) else ""
                if not crm_id:
                    continue
                client_name = row[cols["full name"]] if "full name" in cols and cols["full name"] < len(row) else ""
                paid_ids.append({"crm_id": crm_id, "client_name": client_name, "source": "first_pays"})

    epf = _sheet_by_name(workbook, "epf")
    if epf is not None:
        cols = _header_map(epf)
        if "contact id" not in cols:
            errors.append("EPF tab is missing a 'Contact ID' column.")
        else:
            def cell(row, key):
                idx = cols.get(key)
                return row[idx] if idx is not None and idx < len(row) else None

            for row in epf.iter_rows(min_row=2, values_only=True):
                crm_id = _clean_id(cell(row, "contact id"))
                if not crm_id:
                    continue
                client_name = cell(row, "full name") or ""
                # EPF entries still confirm Cordoba paid us on the client — feeds the
                # CordobaPaidClient ledger / "Cordoba Payout" flag (chargeback gate 2).
                # Unit-only crediting is no longer driven by this tab (owner decision,
                # July 2026): that's now decided directly from the CRM export's own
                # Credit Score column — see crm_parser.py.
                paid_ids.append({"crm_id": crm_id, "client_name": client_name, "source": "epf"})

    chargebacks = _parse_chargebacks(workbook, errors)

    return {"paid_ids": paid_ids, "chargebacks": chargebacks, "errors": errors}
