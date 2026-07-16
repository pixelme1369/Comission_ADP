"""
Parses a historical commission ledger (.xlsx or .csv) handed down from a prior account
manager — NOT a CRM export. One row per client per month already paid (or clawed back) by
that manager, format: Month, ID, Sales Rep, Full Name, Enrolled Debt, To subtract,
Payments Made, Units, Status, Marketing Campaign.

Each row is exactly one of two things (never both, per the source file):
  - Enrolled Debt filled  → a unit the agent was actually paid commission on that month
  - "To subtract" filled  → a clawback dollar amount already deducted from the agent
    that month (the negative number is the deduction; Enrolled Debt is blank on these
    rows since the original enrolled debt isn't repeated here)

There's no Dropped Date, Pay Freq, or payout-date logic to apply here — the prior
manager already resolved which clients got paid and which got clawed back and by how
much, so we just replay those two facts into our own tier math to reconstruct each
(agent, month) period, using the "To subtract" dollar amounts as-is rather than
recomputing them (we don't have enough history to redo that math accurately).

The Month column has no year, so the caller supplies one — the whole file is assumed
to be a single calendar year.
"""

import csv
import io
from collections import defaultdict

from app.calculator import calculate_agent_commission

REQUIRED_COLUMNS = {"month", "id", "sales rep", "enrolled debt", "to subtract"}

MONTH_NUMBERS = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7, "aug": 8,
    "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _clean_id(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    s = str(value).strip()
    # CSV round-trips of Excel data often turn an ID like 1181065497 into "1181065497.0"
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return s


def _parse_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace("$", "").replace(",", ""))
    except ValueError:
        return None


def _parse_int(value) -> int:
    """Lenient int for count-ish columns — bad values become 0 instead of crashing the upload."""
    parsed = _parse_number(value)
    return int(parsed) if parsed is not None else 0


def _header_map_from_row(header_row) -> dict:
    return {str(h).strip().lower(): idx for idx, h in enumerate(header_row) if h and str(h).strip()}


def _read_rows(file_bytes: bytes, filename: str):
    """Returns (cols, data_rows) where data_rows is an iterable of index-able rows,
    for either a .csv or .xlsx file. Raises ValueError on an unreadable file."""
    is_csv = (filename or "").lower().endswith(".csv")

    if is_csv:
        text = file_bytes.decode("utf-8-sig", errors="replace")
        all_rows = list(csv.reader(io.StringIO(text)))
        if not all_rows:
            raise ValueError("File is empty.")
        cols = _header_map_from_row(all_rows[0])
        return cols, all_rows[1:]

    import openpyxl
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        raise ValueError("Could not read the file — expected an .xlsx workbook or .csv file.")
    sheet = workbook[workbook.sheetnames[0]]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    cols = _header_map_from_row(header_row)
    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    return cols, data_rows


def _row_is_blank(row) -> bool:
    return not row or all(v is None or (isinstance(v, str) and not v.strip()) for v in row)


def _zero_unit_result(agent_name: str) -> dict:
    return {
        "agent_name": agent_name,
        "units_cleared": 0,
        "total_cleared_debt": 0.0,
        "cancellation_rate": 0.0,
        "hourly_draw": 0.0,
        "raw_tier": 0,
        "adjusted_tier": 0,
        "tier_rate": 0.0,
        "gross_commission": 0.0,
        "payout": 0.0,
        "payout_type": "none",
        "quality_bonus_eligible": False,
        "cancellation_penalty_applied": False,
        "notes": "Historical import: clawback only, no units cleared this month",
    }


def parse_commission_history(file_bytes: bytes, filename: str, year: int) -> dict:
    """
    Returns {"periods": [{"period_label": "YYYY-MM", "results": [...]}, ...], "errors": [...]}
    Each result dict matches calculate_agent_commission's shape plus clawback_amount,
    net_commission, source, and internal "_cleared_clients"/"_clawback_clients" lists of
    ClientRecord-shaped dicts for the caller to save.
    """
    try:
        cols, data_rows = _read_rows(file_bytes, filename)
    except ValueError as e:
        return {"periods": [], "errors": [str(e)]}

    missing = REQUIRED_COLUMNS - set(cols.keys())
    if missing:
        return {"periods": [], "errors": [f"Missing column(s): {', '.join(sorted(missing))}"]}

    def cell(row, name):
        idx = cols.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    # (period_label, agent_name) -> {"cleared": [...], "clawback": [...]}
    buckets = defaultdict(lambda: {"cleared": [], "clawback": []})
    # crm_id -> enrolled_debt, for cosmetic backfill on clawback rows found later in the file
    debt_by_id = {}
    row_errors = []

    for row_num, row in enumerate(data_rows, start=2):
        if _row_is_blank(row):
            continue

        month_name = str(cell(row, "month") or "").strip().lower()
        month_num = MONTH_NUMBERS.get(month_name)
        crm_id = _clean_id(cell(row, "id"))
        agent_name = str(cell(row, "sales rep") or "").strip()

        if not month_num or not crm_id or not agent_name:
            row_errors.append(f"Row {row_num}: missing Month/ID/Sales Rep — skipped")
            continue

        period_label = f"{year}-{month_num:02d}"
        enrolled_debt = _parse_number(cell(row, "enrolled debt"))
        to_subtract = _parse_number(cell(row, "to subtract"))

        base = {
            "crm_id": crm_id,
            "agent_name": agent_name,
            "client_name": cell(row, "full name"),
            "status": cell(row, "status"),
            "payments_made": _parse_int(cell(row, "payments made")),
        }

        if enrolled_debt is not None:
            base["enrolled_debt"] = enrolled_debt
            debt_by_id[crm_id] = enrolled_debt
            buckets[(period_label, agent_name)]["cleared"].append(base)
        elif to_subtract is not None:
            base["clawback_amount"] = round(abs(to_subtract), 2)
            base["enrolled_debt"] = debt_by_id.get(crm_id, 0.0)
            buckets[(period_label, agent_name)]["clawback"].append(base)
        else:
            row_errors.append(f"Row {row_num} ({agent_name}): neither Enrolled Debt nor "
                               "To subtract is filled — skipped")

    periods = defaultdict(list)  # period_label -> [agent result dicts]

    for (period_label, agent_name), data in buckets.items():
        cleared = data["cleared"]
        clawback = data["clawback"]

        units_cleared = len(cleared)
        total_cleared_debt = sum(c["enrolled_debt"] for c in cleared)
        total_for_rate = units_cleared + len(clawback)
        cancel_rate_pct = (len(clawback) / total_for_rate * 100) if total_for_rate > 0 else 0.0

        if units_cleared > 0:
            result = calculate_agent_commission(
                agent_name=agent_name,
                units_cleared=units_cleared,
                total_cleared_debt=total_cleared_debt,
                cancellation_rate_pct=cancel_rate_pct,
                hourly_draw=0.0,
            )
        else:
            result = _zero_unit_result(agent_name)

        total_clawback = round(sum(c["clawback_amount"] for c in clawback), 2)
        result["clawback_amount"] = total_clawback
        result["net_commission"] = max(0.0, round(result["gross_commission"] - total_clawback, 2))
        result["source"] = "history_import"
        result["pending_units"] = 0
        result["pending_debt"] = 0.0
        result["nsf_flagged"] = False
        result["_cleared_clients"] = cleared
        result["_clawback_clients"] = clawback

        periods[period_label].append(result)

    return {
        "periods": [{"period_label": label, "results": results} for label, results in periods.items()],
        "errors": row_errors,
    }
